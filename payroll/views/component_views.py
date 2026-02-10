"""
component_views.py

This module is used to write methods to the component_urls patterns respectively
"""

import json
import operator
from datetime import datetime
from collections import defaultdict
from datetime import date, datetime, timedelta
from itertools import groupby
from urllib.parse import parse_qs
import csv
# Run - Payroll
import pandas as pd
from django.db import transaction
from django.shortcuts import render


from django.contrib.auth.decorators import login_required, permission_required

import numpy as np
import pandas as pd
from django.apps import apps
from django.contrib import messages
from django.db.models import Sum, Q, F
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, QueryDict
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from django.views.decorators.cache import never_cache
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter

from base.backends import ConfiguredEmailBackend
from base.methods import (
    closest_numbers,
    eval_validate,
    filter_own_records,
    get_key_instances,
    get_next_month_same_date,
    sortby,
)
from base.models import Company
from employee.models import Employee, EmployeeWorkInformation
from horilla.decorators import (
    hx_request_required,
    login_required,
    owner_can_enter,
    permission_required,
)
from horilla.group_by import group_by_queryset
from horilla.horilla_settings import HORILLA_DATE_FORMATS
from horilla.methods import dynamic_attr, get_horilla_model_class, get_urlencode

# from leave.models import AvailableLeave
from notifications.signals import notify
from payroll.filters import (
    AllowanceFilter,
    DeductionFilter,
    LoanAccountFilter,
    PayslipFilter,
    PayslipReGroup,
    ReimbursementFilter,
)
from payroll.forms import component_forms as forms
from payroll.methods.deductions import create_deductions, update_compensation_deduction
from payroll.methods.methods import (
    calculate_employer_contribution,
    compute_net_pay,
    compute_salary_on_period,
    get_daily_salary,
    get_leaves,
    months_between_range,
    paginator_qry,
    save_payslip,
)
from payroll.methods.payslip_calc import (
    calculate_allowance,
    calculate_gross_pay,
    calculate_net_pay_deduction,
    calculate_post_tax_deduction,
    calculate_pre_tax_deduction,
    calculate_tax_deduction,
    calculate_taxable_gross_pay,
)
from payroll.methods.tax_calc import calculate_taxable_amount
from payroll.models.models import (
    Allowance,
    Contract,
    Deduction,
    LoanAccount,
    Payslip,
    Reimbursement,
    ReimbursementMultipleAttachment,
)
from payroll.threadings.mail import MailSendThread
# run payroll
from payroll.forms import PayslipImportForm, EmployeeImportForm
from django.db import transaction
from django.http import HttpResponse
from io import BytesIO, StringIO
from django.contrib.auth.decorators import permission_required
import calendar
from attendance.models import WorkRecords, Attendance
from base.models import Holidays, EmployeeShiftSchedule
from base.methods import get_working_days, get_date_range
from attendance.methods.utils import strtime_seconds

def return_none(a, b):
    return None


operator_mapping = {
    "equal": operator.eq,
    "notequal": operator.ne,
    "lt": operator.lt,
    "gt": operator.gt,
    "le": operator.le,
    "ge": operator.ge,
    "icontains": operator.contains,
    "range": return_none,
}

def payroll_calculation(employee, start_date, end_date, wage=None):
    """
    Calculate payroll components for the specified employee within the given date range.

    Args:
        employee (Employee): The employee for whom the payroll is calculated.
        start_date (date): The start date of the payroll period.
        end_date (date): The end date of the payroll period.
        wage (float, optional): The wage to be used for calculation. Defaults to contract wage.

    Returns:
        dict: A dictionary containing the calculated payroll components.
    """
    # CRITICAL FIX: Use a dict to store total_allowance to completely avoid Python's local variable scoping issues
    # This prevents UnboundLocalError even if Python's bytecode compiler marks it as local
    _vars = {"total_allowance": 0.0}
    
    print("\n========== START PAYROLL CALCULATION ==========")
    print("Employee:", employee, "| Period:", start_date, "→", end_date)

    basic_pay_details = compute_salary_on_period(employee, start_date, end_date, wage=wage)
    if basic_pay_details is None:
        raise ValueError(f"No active contract found for employee {employee} in this period.")

    contract = basic_pay_details["contract"]
    contract_wage = basic_pay_details["contract_wage"]
    basic_pay = basic_pay_details["basic_pay"]
    loss_of_pay = basic_pay_details["loss_of_pay"]
    paid_days = basic_pay_details["paid_days"]
    unpaid_days = basic_pay_details["unpaid_days"]
    month_data_list = basic_pay_details.get("month_data", [])
    # If it's a list, take the first dict, otherwise assume it's already a dict
    if isinstance(month_data_list, list) and month_data_list:
        working_days_details = month_data_list[0]
    else:
        working_days_details = month_data_list


    print("Basic Pay Details:", basic_pay_details)
    print(f"Basic Pay BEFORE update_compensation_deduction: {basic_pay}")

    updated_basic_pay_data = update_compensation_deduction(
        employee, basic_pay, "basic_pay", start_date, end_date
    )
    print(f"Basic Pay AFTER update_compensation_deduction: {updated_basic_pay_data['compensation_amount']}")
    print(f"Basic Pay Deductions Applied: {updated_basic_pay_data['deductions']}")
    if updated_basic_pay_data['deductions']:
        print(f"WARNING: Deductions are being applied to basic_pay! This may reduce the Updated Basic Pay value.")
    basic_pay = updated_basic_pay_data["compensation_amount"]
    basic_pay_deductions = updated_basic_pay_data["deductions"]

    loss_of_pay_amount = 0
    if contract and hasattr(contract, 'deduct_leave_from_basic_pay'):
        if not contract.deduct_leave_from_basic_pay:
            loss_of_pay_amount = loss_of_pay
        else:
            basic_pay = basic_pay - loss_of_pay_amount
    else:
        # If no contract, treat LOP as deduction (standard behavior)
        loss_of_pay_amount = loss_of_pay

    kwargs = {
        "employee": employee,
        "start_date": start_date,
        "end_date": end_date,
        "basic_pay": basic_pay,
        "day_dict": working_days_details,
        "paid_days": paid_days,
        "total_days_in_month": working_days_details.get("total_days_in_month", 30) if isinstance(working_days_details, dict) else 30,
        "contract_wage": contract_wage,
        "payslip_data": {},
    }


    # --- Calculate dynamic allowances ---
    # Use _vars dict to store total_allowance to avoid scoping issues
    allowances = {"allowances": []}
    try:
        calculated_allowances = calculate_allowance(**kwargs)
        if calculated_allowances and isinstance(calculated_allowances, dict) and "allowances" in calculated_allowances:
            allowances = calculated_allowances
            _vars["total_allowance"] = float(sum(allowance.get("amount", 0) for allowance in allowances["allowances"] if allowance and isinstance(allowance, dict)))
        else:
            allowances = {"allowances": []}
            _vars["total_allowance"] = 0.0
    except Exception as e:
        print(f"Error calculating allowances: {e}")
        import traceback
        traceback.print_exc()
        allowances = {"allowances": []}
        _vars["total_allowance"] = 0.0  # Ensure it's set even on exception

    # Create total_allowance variable for backward compatibility
    total_allowance = _vars["total_allowance"]
    kwargs["allowances"] = allowances
    kwargs["total_allowance"] = float(total_allowance)

    print("Dynamic Allowances:", allowances.get("allowances", []))
    print("Total Dynamic Allowance:", total_allowance)

    # --- Calculate gross pay (including fixed allowances from contract) ---
    updated_gross_pay_data = calculate_gross_pay(**kwargs)
    print("DEBUG Gross Pay Data:", updated_gross_pay_data)

    gross_pay = updated_gross_pay_data.get("gross_pay", 0.0)
    housing_allowance = updated_gross_pay_data.get("housing_allowance", 0.0)
    transport_allowance = updated_gross_pay_data.get("transport_allowance", 0.0)
    other_allowance = updated_gross_pay_data.get("other_allowance", 0.0)
    gross_pay_deductions = updated_gross_pay_data.get("deductions", [])

    # --- Add fixed allowances to total allowance ---
    # Use _vars dict to update total_allowance to avoid any scoping issues
    _vars["total_allowance"] = _vars["total_allowance"] + float(housing_allowance) + float(transport_allowance) + float(other_allowance)
    total_allowance = _vars["total_allowance"]  # Update local variable for backward compatibility
    kwargs["total_allowance"] = float(total_allowance)
    kwargs["gross_pay"] = float(gross_pay)

    print("✅ Gross Pay:", gross_pay)
    print("Housing:", housing_allowance, "| Transport:", transport_allowance, "| Other:", other_allowance)
    # --- Deductions ---
    pretax_deductions = calculate_pre_tax_deduction(**kwargs)
    post_tax_deductions = calculate_post_tax_deduction(**kwargs)
    installments = pretax_deductions["installments"] | post_tax_deductions["installments"]

    # taxable_gross_pay = calculate_taxable_gross_pay(**kwargs)
    gross_data = calculate_gross_pay(**kwargs)
    taxable_gross_pay = calculate_taxable_gross_pay(gross_data=gross_data, **kwargs)

    tax_deductions = calculate_tax_deduction(**kwargs)
    federal_tax = calculate_taxable_amount(**kwargs)

    total_pretax_deduction = sum(item["amount"] for item in pretax_deductions["pretax_deductions"])
    total_post_tax_deduction = sum(item["amount"] for item in post_tax_deductions["post_tax_deductions"])
    total_tax_deductions = sum(item["amount"] for item in tax_deductions["tax_deductions"])

    total_deductions = (
        total_pretax_deduction
        + total_post_tax_deduction
        + total_tax_deductions
        + federal_tax
        + loss_of_pay_amount
    )

    # --- Compute Net Pay ---
    net_pay = gross_pay - total_deductions
    net_pay = compute_net_pay(
        net_pay=net_pay,
        gross_pay=gross_pay,
        total_pretax_deduction=total_pretax_deduction,
        total_post_tax_deduction=total_post_tax_deduction,
        total_tax_deductions=total_tax_deductions,
        federal_tax=federal_tax,
        loss_of_pay_amount=loss_of_pay_amount,
        loss_of_pay=loss_of_pay,
    )

    updated_net_pay_data = update_compensation_deduction(
        employee, net_pay, "net_pay", start_date, end_date
    )
    net_pay = updated_net_pay_data["compensation_amount"]
    update_net_pay_deductions = updated_net_pay_data["deductions"]

    net_pay_deductions = calculate_net_pay_deduction(
        net_pay,
        post_tax_deductions["net_pay_deduction"],
        **kwargs,
    )

    net_pay_deduction_list = net_pay_deductions["net_pay_deductions"]
    for deduction in update_net_pay_deductions:
        net_pay_deduction_list.append(deduction)
    net_pay = net_pay - net_pay_deductions["net_deduction"]

    print("✅ Final Net Pay:", net_pay)
    print("========== END PAYROLL CALCULATION ==========\n")

    # --- Final Payslip Data ---
    # Ensure basic_pay is properly stored (it should not be 0 unless actually calculated as 0)
    final_basic_pay = round(float(basic_pay), 2) if basic_pay is not None else 0.0
    print(f"[PAYROLL_CALC] Storing final basic_pay: {final_basic_pay}")
    
    payslip_data = {
        "employee": employee,
        "contract_wage": contract_wage,
        "basic_pay": final_basic_pay,  # Ensure it's stored correctly
        "gross_pay": gross_pay,
        "housing_allowance": housing_allowance,
        "transport_allowance": transport_allowance,
        "other_allowance": other_allowance,
        "taxable_gross_pay": taxable_gross_pay["taxable_gross_pay"]
        if isinstance(taxable_gross_pay, dict)
        else taxable_gross_pay,
        "net_pay": net_pay,
        "allowances": allowances["allowances"],
        "paid_days": paid_days,
        "unpaid_days": unpaid_days,
        "basic_pay_deductions": basic_pay_deductions,
        "gross_pay_deductions": gross_pay_deductions,
        "pretax_deductions": pretax_deductions["pretax_deductions"],
        "post_tax_deductions": post_tax_deductions["post_tax_deductions"],
        "tax_deductions": tax_deductions["tax_deductions"],
        "net_deductions": net_pay_deduction_list,
        "total_deductions": total_deductions,
        "loss_of_pay": loss_of_pay,
        "federal_tax": federal_tax,
        "start_date": start_date,
        "end_date": end_date,
        "range": f"{start_date.strftime('%b %d %Y')} - {end_date.strftime('%b %d %Y')}",
    }

    # --- Convert to JSON for saving ---
    data_to_json = payslip_data.copy()
    data_to_json["employee"] = employee.id
    data_to_json["start_date"] = start_date.strftime("%Y-%m-%d")
    data_to_json["end_date"] = end_date.strftime("%Y-%m-%d")

    # ✅ Keep both formats:
    payslip_data["json_data"] = json.dumps(data_to_json)  # String (for create_payslip)
    payslip_data["pay_head_data"] = data_to_json          # Dict (for view_created_payslip)
    payslip_data["installments"] = installments

    return payslip_data


@login_required
@hx_request_required
def allowances_deductions_tab(request, emp_id):
    """
    Retrieve and render the allowances and deductions applicable to an employee.

    This view function retrieves the active contract, basic pay, allowances, and
    deductions for a specified employee. It filters allowances and deductions
    based on various conditions, including specific employee assignments and
    condition-based rules. The results are then rendered in the allowance and
    deduction tab template.
    """
    employee_deductions = []
    employee_allowances = []
    employee = Employee.objects.get(id=emp_id)
    active_contracts = employee.contract_set.filter(contract_status="active").first()
    basic_pay = active_contracts.wage if active_contracts else None
    if basic_pay:
        allowances = (
            Allowance.objects.filter(specific_employees=employee)
            | Allowance.objects.filter(is_condition_based=True).exclude(
                exclude_employees=employee
            )
            | Allowance.objects.filter(include_active_employees=True).exclude(
                exclude_employees=employee
            )
        )

        for allowance in allowances:
            applicable = True
            if allowance.is_condition_based:
                conditions = list(
                    allowance.other_conditions.values_list(
                        "field", "condition", "value"
                    )
                )
                conditions.append(
                    (
                        allowance.field,
                        allowance.condition,
                        allowance.value.lower().replace(" ", "_"),
                    )
                )
                for field, operator, value in conditions:
                    val = dynamic_attr(employee, field)
                    if val is None or not operator_mapping.get(operator)(
                        val, type(val)(value)
                    ):
                        applicable = False
                        break
            if applicable and allowance not in employee_allowances:
                employee_allowances.append(allowance)

        employee_allowances = [
            allowance
            for allowance in employee_allowances
            if operator_mapping.get(allowance.if_condition)(
                basic_pay if allowance.if_choice == "basic_pay" else 0,
                allowance.if_amount,
            )
        ]

        # Find the applicable deductions for the employee
        deductions = (
            Deduction.objects.filter(
                specific_employees=employee,
            )
            | Deduction.objects.filter(
                is_condition_based=True,
            ).exclude(exclude_employees=employee)
            | Deduction.objects.filter(
                include_active_employees=True,
            ).exclude(exclude_employees=employee)
        )
        for deduction in deductions:
            applicable = True
            if deduction.is_condition_based:
                conditions = list(
                    deduction.other_conditions.values_list(
                        "field", "condition", "value"
                    )
                )
                conditions.append(
                    (
                        deduction.field,
                        deduction.condition,
                        deduction.value.lower().replace(" ", "_"),
                    )
                )
                for field, operator, value in conditions:
                    val = dynamic_attr(employee, field)
                    if val is None or not operator_mapping.get(operator)(
                        val, type(val)(value)
                    ):
                        applicable = False
                        break
            if applicable:
                employee_deductions.append(deduction)

    allowance_ids = (
        json.dumps([instance.id for instance in employee_deductions])
        if employee_deductions
        else None
    )
    deduction_ids = (
        json.dumps([instance.id for instance in employee_deductions])
        if employee_deductions
        else None
    )
    context = {
        "active_contracts": active_contracts,
        "basic_pay": basic_pay,
        "allowances": employee_allowances if employee_allowances else None,
        "allowance_ids": allowance_ids,
        "deductions": employee_deductions if employee_deductions else None,
        "deduction_ids": deduction_ids,
        "employee": employee,
    }
    return render(request, "tabs/allowance_deduction-tab.html", context=context)


@login_required
@permission_required("payroll.add_allowance")
def create_allowance(request):
    """
    This method is used to create allowance condition template
    """
    form = forms.AllowanceForm()
    if request.method == "POST":
        form = forms.AllowanceForm(request.POST)
        if form.is_valid():
            form.save()
            form = forms.AllowanceForm()
            messages.success(request, _("Allowance created."))
            return redirect(view_allowance)
    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@permission_required("payroll.view_allowance")
def view_allowance(request):
    """
    This method is used render template to view all the allowance instances
    """
    allowances = Allowance.objects.exclude(only_show_under_employee=True)
    allowance_filter = AllowanceFilter(request.GET)
    allowances = paginator_qry(allowances, request.GET.get("page"))
    allowance_ids = json.dumps([instance.id for instance in allowances.object_list])
    return render(
        request,
        "payroll/allowance/view_allowance.html",
        {
            "allowances": allowances,
            "f": allowance_filter,
            "allowance_ids": allowance_ids,
        },
    )


@login_required
@hx_request_required
def view_single_allowance(request, allowance_id):
    """
    This method is used render template to view the selected allowance instances
    """
    previous_data = get_urlencode(request)
    allowance = Allowance.find(allowance_id)
    allowance_ids_json = request.GET.get("instances_ids")
    context = {
        "allowance": allowance,
    }
    if allowance_ids_json:
        allowance_ids = json.loads(allowance_ids_json)
        previous_id, next_id = closest_numbers(allowance_ids, allowance_id)
        context["next"] = next_id
        context["previous"] = previous_id
        context["allowance_ids"] = allowance_ids
    context["pd"] = previous_data
    return render(
        request,
        "payroll/allowance/view_single_allowance.html",
        context,
    )


@login_required
@hx_request_required
@permission_required("payroll.view_allowance")
def filter_allowance(request):
    """
    Filter and retrieve a list of allowances based on the provided query parameters.
    """
    query_string = request.GET.urlencode()
    allowances = AllowanceFilter(request.GET).qs.exclude(only_show_under_employee=True)
    list_view = "payroll/allowance/list_allowance.html"
    card_view = "payroll/allowance/card_allowance.html"
    template = card_view
    if request.GET.get("view") == "list":
        template = list_view
    allowances = sortby(request, allowances, "sortby")
    allowances = paginator_qry(allowances, request.GET.get("page"))
    allowance_ids = json.dumps([instance.id for instance in allowances.object_list])
    data_dict = parse_qs(query_string)
    get_key_instances(Allowance, data_dict)
    return render(
        request,
        template,
        {
            "allowances": allowances,
            "pd": query_string,
            "filter_dict": data_dict,
            "allowance_ids": allowance_ids,
        },
    )


@login_required
@permission_required("payroll.change_allowance")
def update_allowance(request, allowance_id, **kwargs):
    """
    This method is used to update the allowance
    Args:
        id : allowance instance id
    """
    instance = Allowance.objects.get(id=allowance_id)
    form = forms.AllowanceForm(instance=instance)
    if request.method == "POST":
        form = forms.AllowanceForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, _("Allowance updated."))
            return redirect(view_allowance)
    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@hx_request_required
@permission_required("payroll.delete_allowance")
def delete_allowance(request, allowance_id):
    """
    This method is used to delete the allowance instance
    """
    previous_data = get_urlencode(request)
    try:
        allowance = Allowance.objects.filter(id=allowance_id).first()
        if allowance:
            allowance.delete()
            messages.success(request, _("Allowance deleted successfully"))
        else:
            messages.error(request, _("Allowance not found"))
    except Exception as e:
        messages.error(request, _("An error occurred while deleting the allowance"))
        messages.error(request, str(e))

    if (
        request.path.split("/")[2] == "delete-employee-allowance"
        or not Allowance.objects.exists()
    ):
        return HttpResponse("<script>window.location.reload();</script>")

    instances_ids = request.GET.get("instances_ids")
    if instances_ids:
        instances_list = json.loads(instances_ids)
        previous_instance, next_instance = closest_numbers(instances_list, allowance_id)
        if allowance_id in instances_list:
            instances_list.remove(allowance_id)
            url = f"/payroll/single-allowance-view/{next_instance}"
            params = f"?{previous_data}&instances_ids={instances_list}"
            return redirect(url + params)

    return redirect(f"/payroll/filter-allowance?{previous_data}")


@login_required
@permission_required("payroll.add_deduction")
def create_deduction(request):
    """
    This method is used to create deduction
    """
    form = forms.DeductionForm()
    if request.method == "POST":
        form = forms.DeductionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _("Deduction created."))
            return redirect(view_deduction)
    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@permission_required("payroll.view_allowance")
def view_deduction(request):
    """
    This method is used render template to view all the deduction instances
    """

    deductions = Deduction.objects.exclude(only_show_under_employee=True)
    deduction_filter = DeductionFilter(request.GET)
    deductions = paginator_qry(deductions, request.GET.get("page"))
    deduction_ids = json.dumps([instance.id for instance in deductions.object_list])
    return render(
        request,
        "payroll/deduction/view_deduction.html",
        {
            "deductions": deductions,
            "f": deduction_filter,
            "deduction_ids": deduction_ids,
        },
    )


@login_required
@hx_request_required
def view_single_deduction(request, deduction_id):
    """
    Render template to view a single deduction instance with navigation.
    """
    previous_data = get_urlencode(request)
    deduction = Deduction.objects.filter(id=deduction_id).first()
    context = {"deduction": deduction, "pd": previous_data}

    # Handle deduction IDs and navigation
    deduction_ids_json = request.GET.get("instances_ids")
    if deduction_ids_json:
        deduction_ids = json.loads(deduction_ids_json)
        context["previous"], context["next"] = closest_numbers(
            deduction_ids, deduction_id
        )
        context["deduction_ids"] = deduction_ids

    # Determine htmx load URL and target
    HTTP_REFERER = request.META.get("HTTP_REFERER", "")
    referer_parts = HTTP_REFERER.rstrip("/").split("/")

    if "view-deduction" in referer_parts:
        context.update(
            {
                "load_hx_url": f"/payroll/filter-deduction?{previous_data}",
                "load_hx_target": "#payroll-deduction-container",
            }
        )
    elif referer_parts[-2:] == ["employee-view", str(referer_parts[-1])]:
        try:
            context.update(
                {
                    "load_hx_url": f"/payroll/allowances-deductions-tab/{int(referer_parts[-1])}",
                    "load_hx_target": "#allowance_deduction",
                }
            )
        except ValueError:
            pass
    elif HTTP_REFERER.endswith("employee-profile/"):
        context.update(
            {
                "load_hx_url": f"/payroll/allowances-deductions-tab/{request.user.employee_get.id}",
                "load_hx_target": "#allowance_deduction",
            }
        )
    else:
        context.update({"load_hx_url": None, "load_hx_target": None})

    return render(request, "payroll/deduction/view_single_deduction.html", context)


@login_required
@hx_request_required
@permission_required("payroll.view_allowance")
def filter_deduction(request):
    """
    This method is used search the deduction
    """
    query_string = request.GET.urlencode()
    deductions = DeductionFilter(request.GET).qs.exclude(only_show_under_employee=True)
    list_view = "payroll/deduction/list_deduction.html"
    card_view = "payroll/deduction/card_deduction.html"
    template = card_view
    if request.GET.get("view") == "list":
        template = list_view
    deductions = sortby(request, deductions, "sortby")
    deductions = paginator_qry(deductions, request.GET.get("page"))
    deduction_ids = json.dumps([instance.id for instance in deductions.object_list])
    data_dict = parse_qs(query_string)
    get_key_instances(Deduction, data_dict)
    return render(
        request,
        template,
        {
            "deductions": deductions,
            "pd": query_string,
            "filter_dict": data_dict,
            "deduction_ids": deduction_ids,
        },
    )


@login_required
@permission_required("payroll.change_deduction")
def update_deduction(request, deduction_id, **kwargs):
    """
    This method is used to update the deduction instance
    """
    instance = Deduction.objects.get(id=deduction_id)
    form = forms.DeductionForm(instance=instance)
    if request.method == "POST":
        form = forms.DeductionForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, _("Deduction updated."))
            return redirect(view_deduction)
    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@hx_request_required
@permission_required("payroll.delete_deduction")
def delete_deduction(request, deduction_id, emp_id=None):
    instances_ids = request.GET.get("instances_ids")
    next_instance = None
    instances_list = None
    previous_data = ""
    if instances_ids:
        previous_data = get_urlencode(request)
        instances_list = json.loads(instances_ids)
        previous_instance, next_instance = closest_numbers(instances_list, deduction_id)
        instances_list.remove(deduction_id)
    deduction = Deduction.objects.filter(id=deduction_id).first()
    if deduction:
        deduction.delete()
        messages.success(request, _("Deduction deleted successfully"))
    else:
        messages.error(request, _("Deduction not found"))

    paths = {
        "payroll-deduction-container": f"/payroll/filter-deduction?{request.GET.urlencode()}",
        "allowance_deduction": f"/payroll/allowances-deductions-tab/{emp_id}",
        "objectDetailsModalTarget": f"/payroll/single-deduction-view/{next_instance}?{previous_data}&instances_ids={instances_list}",
    }
    http_hx_target = request.META.get("HTTP_HX_TARGET")
    redirected_path = paths.get(http_hx_target)
    if http_hx_target:
        if (
            http_hx_target == "payroll-deduction-container"
            and not Deduction.objects.filter()
        ):
            return HttpResponse("<script>window.location.reload();</script>")
        if redirected_path:
            return redirect(redirected_path)
    default_redirect = (
        request.path if http_hx_target else request.META.get("HTTP_REFERER", "/")
    )
    return HttpResponseRedirect(default_redirect)


def get_month_start_end(year):
    start_end_dates = []
    for month in range(1, 13):
        # Start date is the first day of the month
        start_date = date(year, month, 1)

        # Calculate the last day of the month
        if month == 12:  # December
            end_date = date(year, 12, 31)
        else:
            next_month = date(year, month + 1, 1)
            end_date = next_month - timedelta(days=1)

        start_end_dates.append((start_date, end_date))
    return start_end_dates


@login_required
@permission_required("payroll.add_payslip")
def generate_payslip(request):
    """
    Generate payslips for selected employees within a specified date range.

    Requires the user to be logged in and have the 'payroll.add_payslip' permission.

    """
    if (
        request.META.get("HTTP_HX_REQUEST")
        and request.META.get("HTTP_HX_TARGET") == "objectCreateModalTarget"
    ):
        bulk_form = forms.GeneratePayslipForm()
        return render(
            request,
            "payroll/payslip/bulk_create_payslip.html",
            {"bulk_form": bulk_form},
        )
    payslips = []
    json_data = []
    form = forms.GeneratePayslipForm()
    if request.method == "POST":
        form = forms.GeneratePayslipForm(request.POST)
        if form.is_valid():
            instances = []
            employees = form.cleaned_data["employee_id"]
            start_date = form.cleaned_data["start_date"]
            end_date = form.cleaned_data["end_date"]

            group_name = form.cleaned_data["group_name"]
            for employee in employees:
                contract = Contract.objects.filter(
                    employee_id=employee, contract_status="active"
                ).first()
                if start_date < contract.contract_start_date:
                    start_date = contract.contract_start_date
                payslip = payroll_calculation(employee, start_date, end_date)
                payslips.append(payslip)
                json_data.append(payslip["json_data"])

                payslip["payslip"] = payslip
                data = {}
                data["employee"] = employee
                data["group_name"] = group_name
                data["start_date"] = payslip["start_date"]
                data["end_date"] = payslip["end_date"]
                data["status"] = "draft"
                data["contract_wage"] = payslip["contract_wage"]
                data["basic_pay"] = payslip["basic_pay"]
                data["gross_pay"] = payslip["gross_pay"]
                data["deduction"] = payslip["total_deductions"]
                data["net_pay"] = payslip["net_pay"]
                data["pay_data"] = json.loads(payslip["json_data"])
                calculate_employer_contribution(data)
                data["installments"] = payslip["installments"]
                instance = save_payslip(**data)
                instances.append(instance)
                notify.send(
                    request.user.employee_get,
                    recipient=employee.employee_user_id,
                    verb="Payslip has been generated for you.",
                    verb_ar="تم إصدار كشف راتب لك.",
                    verb_de="Gehaltsabrechnung wurde für Sie erstellt.",
                    verb_es="Se ha generado la nómina para usted.",
                    verb_fr="La fiche de paie a été générée pour vous.",
                    redirect=reverse(
                        "view-created-payslip", kwargs={"payslip_id": instance.id}
                    ),
                    icon="close",
                )
            messages.success(request, f"{employees.count()} payslip saved as draft")
            return redirect(
                f"/payroll/view-payslip?group_by=group_name&active_group={group_name}"
            )

    return render(request, "payroll/common/form.html", {"form": form})


@login_required
@hx_request_required
def check_contract_start_date(request):
    """
    Check if the employee's contract start date is after the provided payslip start date.
    """
    employee_id = request.GET.get("employee_id")
    start_date = request.GET.get("start_date")

    contract = Contract.objects.filter(
        employee_id=employee_id, contract_status="active"
    ).first()

    if not contract or start_date >= str(contract.contract_start_date):
        return HttpResponse("")

    title_message = _(
        "When this payslip is run, the payslip start date will be updated to match the employee contract start date."
    )
    text_content = _("Employee Contract Start Date")

    return HttpResponse(
        format_html(
            """
        <div id='messageDiv' style='background-color: hsl(48, 100%, 94%);
            border: 1px solid hsl(46, 97%, 88%);
            border-radius: 18px; padding:5px; font-weight: bold; display: flex;'>
            {text_content}: {contract_start_date}
            <img style='width: 20px; height: 20px; cursor: pointer;'
                src='/static/images/ui/info.png' class='ml-2' title='{title_message}'>
        </div>
        """,
            text_content=text_content,
            contract_start_date=contract.contract_start_date,
            title_message=title_message,
        )
    )


@login_required
@permission_required("payroll.add_payslip")
def create_payslip(request, new_post_data=None):
    """
    Create or Edit a payslip for an employee.
    Args:
        request: The HTTP request object.
    Returns:
        A rendered HTML template for the payslip creation/editing form.
    """
    if new_post_data:
        request.POST = new_post_data
    
    # --- Edit Mode: Check for payslip_id in GET (to load form) ---
    payslip_id = request.GET.get("payslip_id")
    payslip_instance = None
    if payslip_id:
        payslip_instance = Payslip.objects.filter(id=payslip_id).first()

    form = forms.PayslipForm()

    if request.method == "POST":
        # --- Edit Mode: Save Changes ---
        payslip_id = request.POST.get("payslip_id")
        if payslip_id:
            instance = Payslip.objects.get(id=payslip_id)
            
            # Update direct fields
            instance.status = request.POST.get("status")
            instance.start_date = request.POST.get("start_date")
            instance.end_date = request.POST.get("end_date")
            instance.basic_pay = float(request.POST.get("basic_pay") or 0)
            instance.gross_pay = float(request.POST.get("gross_pay") or 0)
            instance.net_pay = float(request.POST.get("net_pay") or 0)
            instance.deduction = float(request.POST.get("deduction") or 0)

            # Update JSON data (pay_head_data)
            pay_head_data = instance.pay_head_data or {}
            
            # Helper to update safe floats
            def get_float(key):
                return float(request.POST.get(key) or 0)

            # Allowances
            pay_head_data["housing_allowance"] = get_float("housing_allowance")
            pay_head_data["transport_allowance"] = get_float("transport_allowance")
            pay_head_data["other_allowance"] = get_float("other_allowance")
            
            # Other components
            pay_head_data["loss_of_pay"] = get_float("loss_of_pay")
            pay_head_data["overtime"] = get_float("overtime")
            pay_head_data["salary_advance_loan_recovery"] = get_float("salary_advance_loan_recovery")
            pay_head_data["salary_advance"] = get_float("salary_advance")
            pay_head_data["bonus"] = get_float("bonus")

            instance.pay_head_data = pay_head_data
            instance.save()
            
            messages.success(request, _("Payslip Updated Successfully"))
            return HttpResponse(
                f'<script>window.location.reload()</script>'
            )

        # --- Create Mode: Simplified Horilla-style logic ---
        employee_id = request.POST.get("employee_id")
        start_date = (
            datetime.strptime(request.POST.get("start_date"), "%Y-%m-%d").date()
            if isinstance(request.POST.get("start_date"), str)
            else request.POST.get("start_date")
        )

        if employee_id and start_date:
            contract = Contract.objects.filter(
                employee_id=employee_id, contract_status="active"
            ).first()

            if contract and start_date < contract.contract_start_date:
                new_post_data = request.POST.copy()
                new_post_data["start_date"] = contract.contract_start_date
                request.POST = new_post_data
        
        form = forms.PayslipForm(request.POST)
        if form.is_valid():
            employee = form.cleaned_data["employee_id"]
            start_date = form.cleaned_data["start_date"]
            end_date = form.cleaned_data["end_date"]
            payslip = Payslip.objects.filter(
                employee_id=employee, start_date=start_date, end_date=end_date
            ).first()

            payslip_data = payroll_calculation(employee, start_date, end_date)
            payslip_data["payslip"] = payslip
            data = {}
            data["employee"] = employee
            data["start_date"] = payslip_data["start_date"]
            data["end_date"] = payslip_data["end_date"]
            data["status"] = (
                "draft"
                if request.GET.get("status") is None
                else request.GET["status"]
            )
            data["contract_wage"] = payslip_data["contract_wage"]
            data["basic_pay"] = payslip_data["basic_pay"]
            data["gross_pay"] = payslip_data["gross_pay"]
            data["deduction"] = payslip_data["total_deductions"]
            data["net_pay"] = payslip_data["net_pay"]
            data["pay_data"] = json.loads(payslip_data["json_data"])
            calculate_employer_contribution(data)
            data["installments"] = payslip_data["installments"]
            payslip_data["instance"] = save_payslip(**data)
            form = forms.PayslipForm()
            messages.success(request, _("Payslip Saved"))
            payslip = payslip_data["instance"]
            notify.send(
                request.user.employee_get,
                recipient=employee.employee_user_id,
                verb="Payslip has been generated for you.",
                verb_ar="تم إصدار كشف راتب لك.",
                verb_de="Gehaltsabrechnung wurde für Sie erstellt.",
                verb_es="Se ha generado la nómina para usted.",
                verb_fr="La fiche de paie a été générée pour vous.",
                redirect=reverse(
                    "view-created-payslip", kwargs={"payslip_id": payslip.pk}
                ),
                icon="close",
            )
            return HttpResponse(
                f'<script>window.location.href = "/payroll/view-payslip/{payslip_data["instance"].id}/"</script>'
            )
    
    # Prepare context for Edit Mode
    context = {"individual_form": form}
    if payslip_instance:
        context["payslip"] = payslip_instance
        # Extract allowances/components for template usage
        data = payslip_instance.pay_head_data or {}
        context["pay_data"] = {
            "housing_allowance": data.get("housing_allowance", 0),
            "transport_allowance": data.get("transport_allowance", 0),
            "other_allowance": data.get("other_allowance", 0),
            "loss_of_pay": data.get("loss_of_pay", 0),
            "overtime": data.get("overtime", 0),
            "salary_advance_loan_recovery": data.get("salary_advance_loan_recovery", 0),
            "salary_advance": data.get("salary_advance", 0),
            "bonus": data.get("bonus", 0),
        }

    return render(
        request,
        "payroll/payslip/create_payslip.html",
        context,
    )


@login_required
@permission_required("payroll.add_payslip")
def validate_start_date(request):
    """
    This method to validate the contract start date and the pay period start date
    """
    end_datetime = None
    start_datetime = None
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    employee_id = request.GET.getlist("employee_id")
    if start_date:
        start_datetime = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end_datetime = datetime.strptime(end_date, "%Y-%m-%d").date()
    error_message = ""
    response = {"valid": True, "message": error_message}
    for emp_id in employee_id:
        contract = Contract.objects.filter(
            employee_id__id=emp_id, contract_status="active"
        ).first()

        if start_datetime is not None and start_datetime < contract.contract_start_date:
            error_message = f"<ul class='errorlist'><li>The {contract.employee_id}'s \
                contract start date is smaller than pay period start date</li></ul>"
            response["message"] = error_message
            response["valid"] = False

    if (
        start_datetime is not None
        and end_datetime is not None
        and start_datetime > end_datetime
    ):
        error_message = "<ul class='errorlist'><li>The end date must be greater than \
                or equal to the start date.</li></ul>"
        response["message"] = error_message
        response["valid"] = False

    if end_datetime is not None:
        if end_datetime > datetime.today().date():
            error_message = '<ul class="errorlist"><li>The end date cannot be in the future123.</li></ul>'
            response["message"] = error_message
            response["valid"] = False
    return JsonResponse(response)


@login_required
@permission_required("payroll.view_payslip")
def view_individual_payslip(request, employee_id, start_date, end_date):
    """
    This method is used to render the template for viewing a payslip.
    """

    payslip_data = payroll_calculation(employee_id, start_date, end_date)
    return render(
        request,
        "payroll/payslip/individual_payslip.html",
        payslip_data,
    )


@login_required
@never_cache
def view_payslip(request):
    """
    This method is used to render the template for viewing a payslip.
    """
    if request.user.has_perm("payroll.view_payslip"):
        payslips = Payslip.objects.all()
    else:
        payslips = Payslip.objects.filter(employee_id__employee_user_id=request.user)
    export_column = forms.PayslipExportColumnForm()
    filter_form = PayslipFilter(request.GET, payslips)
    payslips = filter_form.qs
    bulk_form = forms.GeneratePayslipForm()
    field = request.GET.get("group_by")
    if field in Payslip.__dict__.keys():
        payslips = payslips.filter(group_name__isnull=False).order_by(field)
    payslips = paginator_qry(payslips, request.GET.get("page"))

    # Normalize pay_head_data for template safety
    for ps in payslips:
        phd = ps.pay_head_data or {}
        if isinstance(phd, str):
            try:
                phd = json.loads(phd)
            except Exception:
                phd = {}
        if not isinstance(phd, dict):
            phd = {}
        if 'allowances' not in phd or phd.get('allowances') is None:
            phd['allowances'] = []
        if 'pretax_deductions' not in phd or phd.get('pretax_deductions') is None:
            phd['pretax_deductions'] = []
        if 'posttax_deductions' not in phd or phd.get('posttax_deductions') is None:
            phd['posttax_deductions'] = []
        ps.pay_head_data = phd
    previous_data = request.GET.urlencode()
    data_dict = parse_qs(previous_data)
    get_key_instances(Payslip, data_dict)
    return render(
        request,
        "payroll/payslip/view_payslips.html",
        {
            "payslips": payslips,
            "f": filter_form,
            "export_column": export_column,
            "export_filter": PayslipFilter(request.GET),
            "bulk_form": bulk_form,
            "filter_dict": data_dict,
            "gp_fields": PayslipReGroup.fields,
        },
    )


@login_required
@hx_request_required
def filter_payslip(request):
    """
    Filter and retrieve a list of payslips based on the provided query parameters.
    """
    query_string = request.GET.urlencode()
    if request.user.has_perm("payroll.view_payslip"):
        payslips = PayslipFilter(request.GET).qs
    else:
        emp_request = request.GET.copy()
        employee = Employee.objects.filter(employee_user_id=request.user.id).first()
        employee_id = employee.id
        emp_request["employee_id"] = str(employee_id)
        payslips = PayslipFilter(emp_request).qs
    template = "payroll/payslip/payslip_table.html"
    view = request.GET.get("view")
    if view == "card":
        template = "payroll/payslip/group_payslips.html"
        payslips = payslips.filter(group_name__isnull=False).order_by("-group_name")
    payslips = sortby(request, payslips, "sortby")
    data_dict = []
    if not request.GET.get("dashboard"):
        data_dict = parse_qs(query_string)
        get_key_instances(Payslip, data_dict)
    if "status" in data_dict:
        status_list = data_dict["status"]
        if len(status_list) > 1:
            data_dict["status"] = [status_list[-1]]
    field = request.GET.get("field")
    if field != "" and field is not None:
        payslips = group_by_queryset(payslips, field, request.GET.get("page"), "page")
        template = "payroll/payslip/group_by.html"
    else:
        payslips = paginator_qry(payslips, request.GET.get("page"))
    
    # Process payslips to ensure paid_days and unpaid_days are calculated correctly
    # This ensures consistency with individual payslip view
    
    
    # Handle pagination - payslips might be a Page object
    payslips_list = payslips
    if hasattr(payslips, 'object_list'):
        payslips_list = payslips.object_list
    elif hasattr(payslips, '__iter__'):
        payslips_list = list(payslips)
    else:
        payslips_list = [payslips] if payslips else []
    
    processed_payslips = []
    for payslip in payslips_list:
        # CRITICAL: Use shared calculation function to ensure consistency with export
        # This ensures table view shows the same data as export (including database allowances/deductions)
        calculated_values = calculate_payslip_values(payslip)
        if calculated_values is None:
            # Skip if calculation fails (e.g., no employee)
            continue
        
        # Values are already updated in payslip.pay_head_data and payslip attributes by calculate_payslip_values
        processed_payslips.append(payslip)
    
    # If payslips was a paginator Page object, update its object_list
    if hasattr(payslips, 'object_list'):
        payslips.object_list = processed_payslips
        final_payslips = payslips
    else:
        final_payslips = processed_payslips if processed_payslips else payslips
    
    return render(
        request,
        template,
        {
            "payslips": final_payslips,
            "pd": query_string,
            "filter_dict": data_dict,
        },
    )



@login_required
@permission_required("payroll.change_payslip")
def payslip_export(request):
    """
    Exports payslip data (with full pay_head_data info and employee details).
    """
    # --- Handle HTMX filter view ---
    if request.META.get("HTTP_HX_REQUEST"):
        return render(
            request,
            "payroll/payslip/payslip_export_filter.html",
            {
                "export_column": forms.PayslipExportColumnForm(),
                "export_filter": PayslipFilter(request.GET),
            },
        )

    # --- Status Choices ---
    choices_mapping = {
        "draft": _("Draft"),
        "review_ongoing": _("Review Ongoing"),
        "confirmed": _("Confirmed"),
        "paid": _("Paid"),
    }

    payslips = PayslipFilter(request.GET).qs
    today_date = date.today().strftime("%Y-%m-%d")
    file_name = f"Payslip_excel_{today_date}.xlsx"

    selected_fields = request.GET.getlist("selected_fields")
    form = forms.PayslipExportColumnForm()

    # --- Handle selected IDs if passed from frontend ---
    if not selected_fields:
        selected_fields = form.fields["selected_fields"].initial
        ids = request.GET.get("ids")
        if ids:
            id_list = json.loads(ids)
            payslips = Payslip.objects.filter(id__in=id_list)

    export_rows = []
    print(f"🟢 Found {payslips.count()} payslips to export")

    # --- Loop through payslips ---
    for ps in payslips:
        print(ps)
        payslip = Payslip.objects.filter(id=ps.id).first()
        print(f"🔵 Processing Payslip ID: {ps}")
        if not payslip:
            print(f"⚠ Payslip not found for ID: {ps.id}")
            continue

        employee = payslip.employee_id
        if not employee:
            print(f"⚠ No employee linked for payslip {payslip.id}")
            continue

        # --- Load pay_head_data safely ---
        data = payslip.pay_head_data or {}
        if isinstance(data, str):
            try:
                data = json.loads(data)
                print(f"✅ Loaded pay_head_data for {employee.badge_id}")
            except Exception as e:
                print(f"⚠ Error loading pay_head_data for {employee.badge_id}: {e}")
                data = {}

        # --- Employee info ---
        print(f"➡ Processing {employee.employee_first_name} ({employee.badge_id})")
        emp_info = EmployeeWorkInformation.objects.filter(employee_id=employee).first()

        emp_name = f"{employee.employee_first_name} {employee.employee_last_name}"
        emp_code = employee.badge_id or ""
        dept_name = emp_info.department_id.department if emp_info and emp_info.department_id else ""
        batch_name = getattr(payslip, "group_name", "") or getattr(getattr(payslip, "batch", None), "name", "")
        join_date = (
            emp_info.date_joining.strftime("%b. %d, %Y")
            if emp_info and emp_info.date_joining
            else ""
        )

        # --- Dates ---
        start_date = payslip.start_date.strftime("%b. %d, %Y") if payslip.start_date else ""
        end_date = payslip.end_date.strftime("%b. %d, %Y") if payslip.end_date else ""

        # --- Pay values ---
        basic_pay = round(float(data.get("basic_pay", payslip.basic_pay or 0)), 2)
        gross_pay = round(float(data.get("gross_pay", payslip.gross_pay or 0)), 2)
        net_pay = round(float(data.get("net_pay", payslip.net_pay or 0)), 2)
        
        # --- MISSING COLUMN: Total Paid days ---
        paid_days = round(float(data.get("paid_days", 0)), 2)

        # --- ✅ Improved Allowances Logic (FIXED: Fetching top-level keys) ---
        # 1. Fetch allowances directly from top-level keys in 'data'
        housing_allowance = round(float(data.get("housing_allowance", 0)), 2)
        transport_allowance = round(float(data.get("transport_allowance", 0)), 2)
        other_allowance = round(float(data.get("other_allowance", 0)), 2)
        
        total_allowance = 0 
        
        # --- Helper recursive function to extract allowances from nested JSON ---
        def extract_allowances(obj):
            nonlocal housing_allowance, transport_allowance, other_allowance

            if isinstance(obj, list):
                for item in obj:
                    extract_allowances(item)

            elif isinstance(obj, dict):
                name = str(obj.get("name", "")).lower()
                amount = float(obj.get("amount", obj.get("value", 0) or 0))

                if amount:
                    if any(x in name for x in ["housing", "housing", "hra", "rent"]):
                        housing_allowance += amount
                    elif any(x in name for x in ["transport", "conveyance"]):
                        transport_allowance += amount
                    elif "other" in name:
                        other_allowance += amount
                    elif any(x in name for x in ["allowance", "bonus"]):
                        other_allowance += amount
                    
                for v in obj.values():
                    extract_allowances(v)

        # 2. Run recursive search only on the 'allowances' list for nested items
        extract_allowances(data.get("allowances", []))

        # 3. Calculate Total Allowance as the sum of its categorized components for consistency
        total_allowance = housing_allowance + transport_allowance + other_allowance

        if total_allowance == 0:
            print(f"⚠ No allowance values found for {emp_code}")
        else:
            print(
                f"✅ Allowances for {emp_code}: Housing={housing_allowance}, Transport={transport_allowance}, Other={other_allowance}, Total={total_allowance}"
            )


        # --- Deductions ---
        # MISSING COLUMN: Other Deductions
        other_deductions_amount = 0
        deduction_keys = [
            "gross_pay_deductions",
            "basic_pay_deductions",
            "pretax_deductions",
            "post_tax_deductions",
            "tax_deductions",
            "net_deductions",
        ]

        # Calculate 'Other Deductions' (sum of all list deductions)
        for key in deduction_keys:
            for d in data.get(key, []):
                other_deductions_amount += float(d.get("amount", 0))
        
        other_deductions_amount = round(other_deductions_amount, 2)
        
        # MISSING COLUMN: LOP (Loss of Pay) is a separate top-level key
        loss_of_pay = round(float(data.get("loss_of_pay", 0)), 2)
        
        # Extract variable components
        overtime = round(float(data.get("overtime", 0) or 0), 2)
        salary_advance = round(float(data.get("salary_advance", 0) or 0), 2)
        bonus = round(float(data.get("bonus", 0) or 0), 2)
        salary_advance_loan_recovery = round(float(data.get("salary_advance_loan_recovery", 0) or 0), 2)
        deduction = round(float(data.get("deduction", payslip.deduction or 0) or 0), 2)
        
        # Total Deductions is the sum of all list deductions (other) and LOP
        total_deductions = round(other_deductions_amount + loss_of_pay, 2)

        # --- Status ---
        status_display = choices_mapping.get(payslip.status, payslip.status)

        # --- Final Row: Including all fixed and new columns ---
        export_rows.append({
            "Employee Id": emp_code,
            "Employee Name": emp_name,
            "Department": dept_name,
            "Basic Pay": basic_pay,
            "Housing Allowance": housing_allowance,
            "Transport Allowance": transport_allowance,
            "Other Allowance": other_allowance,
            "Gross Pay": gross_pay,
            # Variable components
            "Total Paid Days": paid_days,
            "LOP": loss_of_pay,
            "Overtime": overtime,
            "salary_advance_loan_recovery": salary_advance_loan_recovery,
            "Deduction": deduction,
            "salary_advance": salary_advance,
            "bonus": bonus,
            "Net Pay": net_pay,
            # Additional fields
            # "Batch": batch_name,
            # "Employment Start Date": join_date,
            # "Start Date": start_date,
            # "End Date": end_date,
            # "Total Allowances": round(total_allowance, 2),
            # "Other Deductions": other_deductions_amount,
            # "Total Deductions": total_deductions,
            # "Status": status_display,
        })

    # --- Export to Excel ---
    df = pd.DataFrame(export_rows)
    print(f"✅ Prepared {len(df)} rows for Excel export")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    writer = pd.ExcelWriter(response, engine="xlsxwriter")
    df.to_excel(writer, index=False, sheet_name="Payslips", startrow=0)

    workbook = writer.book
    worksheet = writer.sheets["Payslips"]

    # --- Excel Formatting Code with Color Styling (matching payslip_import_download) ---
    # Static components header format (green)
    static_header_fmt = workbook.add_format({
        'bold': True, 
        'bg_color': '#D8E4BC',  # Light green
        'border': 1,
        'align': 'center'
    })
    
    # Variable components header format (orange)
    variable_header_fmt = workbook.add_format({
        'bold': True, 
        'bg_color': '#FCD5B4',  # Orange
        'border': 1,
        'align': 'center'
    })
    
    # Static components data format (green background)
    static_data_fmt = workbook.add_format({
        'bg_color': '#D8E4BC',  # Light green
        'border': 1
    })
    
    # Variable components data format (orange background)
    variable_data_fmt = workbook.add_format({
        'bg_color': '#FCD5B4',  # Orange
        'border': 1
    })
    
    # Default format for additional columns
    default_header_fmt = workbook.add_format({
        'bold': True, 
        'bg_color': '#D3D3D3',  # Gray
        'border': 1,
        'align': 'center'
    })
    
    # Define static and variable columns (matching payslip_import_download)
    static_cols = ["Employee Id", "Employee Name", "Department", "Basic Pay", "Housing Allowance", "Transport Allowance", "Other Allowance", "Gross Pay"]
    variable_cols = ["Total Paid Days", "LOP", "Overtime", "salary_advance_loan_recovery", "Deduction", "salary_advance", "bonus", "Net Pay"]

    # Write headers with appropriate formatting
    for col_num, col_name in enumerate(df.columns.values):
        if col_name in static_cols:
            worksheet.write(0, col_num, col_name, static_header_fmt)
        elif col_name in variable_cols:
            worksheet.write(0, col_num, col_name, variable_header_fmt)
        else:
            worksheet.write(0, col_num, col_name, default_header_fmt)
        
        # Set column width
        try:
            column_len = max(df[col_name].astype(str).map(len).max(), len(col_name)) + 2
        except Exception:
            column_len = len(col_name) + 2
        worksheet.set_column(col_num, col_num, column_len)
    
    # Apply row colors for data rows - only up to rows where values are present
    # Row 0 is header, so data starts from row 1
    for row_idx in range(1, len(df) + 1):
        for col_num, col_name in enumerate(df.columns.values):
            # Get cell value from dataframe
            cell_value = df.iloc[row_idx - 1, col_num]
            
            # Apply formatting based on column type, preserving the value
            if col_name in static_cols:
                # Static columns - green background
                if pd.notna(cell_value):
                    worksheet.write(row_idx, col_num, cell_value, static_data_fmt)
                else:
                    worksheet.write(row_idx, col_num, "", static_data_fmt)
            elif col_name in variable_cols:
                # Variable columns - orange background
                if pd.notna(cell_value):
                    worksheet.write(row_idx, col_num, cell_value, variable_data_fmt)
                else:
                    worksheet.write(row_idx, col_num, "", variable_data_fmt)

    worksheet.set_row(0, 25)
    writer.close()
    return response



@login_required
@permission_required("payroll.add_allowance")
def hx_create_allowance(request):
    """
    This method is used to render htmx allowance form
    """
    form = forms.AllowanceForm()
    return render(request, "payroll/htmx/form.html", {"form": form})


@login_required
@permission_required("payroll.add_payslip")
def send_slip(request):
    """
    Send payslip method
    """
    email_backend = ConfiguredEmailBackend()
    view = request.GET.get("view")
    payslip_ids = request.GET.getlist("id")
    payslips = Payslip.objects.filter(id__in=payslip_ids)
    if not getattr(
        email_backend, "dynamic_from_email_with_display_name", None
    ) or not len(email_backend.dynamic_from_email_with_display_name):
        messages.error(request, "Email server is not configured")
        if view:
            return HttpResponse("<script>window.location.reload()</script>")
        else:
            return redirect(filter_payslip)

    result_dict = defaultdict(
        lambda: {"employee_id": None, "instances": [], "count": 0}
    )
    for payslip in payslips:
        employee_id = payslip.employee_id
        result_dict[employee_id]["employee_id"] = employee_id
        result_dict[employee_id]["instances"].append(payslip)
        result_dict[employee_id]["count"] += 1
    mail_thread = MailSendThread(request, result_dict=result_dict, ids=payslip_ids)
    mail_thread.start()
    messages.info(request, "Mail processing")
    if view:
        return HttpResponse("<script>window.location.reload()</script>")
    else:
        return redirect(filter_payslip)


# @login_required
# @permission_required("payroll.add_allowance")
# def add_bonus(request):
#     employee_id = request.GET["employee_id"]
#     payslip_id = request.GET.get("payslip_id")
#     instance = None
#     if payslip_id != "None" and payslip_id:
#         try:
#             instance = Payslip.objects.get(id=payslip_id)
#         except Payslip.DoesNotExist:
#             messages.error(request, _("Payslip not found."))
#             return HttpResponse("<script>window.location.reload()</script>")
#         form = forms.PayslipAllowanceForm(
#             initial={"employee_id": employee_id, "date": instance.start_date}
#         )
#     else:
#         form = forms.BonusForm(initial={"employee_id": employee_id})
#     if request.method == "POST":
#         form = forms.BonusForm(request.POST, initial={"employee_id": employee_id})
#         contract = Contract.objects.filter(
#             employee_id=employee_id, contract_status="active"
#         ).first()
#         employee = Employee.objects.filter(id=employee_id).first()
#         if form.is_valid():
#             form.save()
#             messages.success(request, _("Bonus Added"))
#             if payslip_id != "None" and payslip_id and instance:
#                 # Recalculate payslip in place without deleting it
#                 try:
#                     employee = instance.employee_id
#                     start_date = instance.start_date
#                     end_date = instance.end_date
                    
#                     # Try to recalculate using payroll_calculation or calculate_payslip_from_attendance
#                     contract = Contract.objects.filter(
#                         employee_id=employee.id, contract_status="active"
#                     ).first()
                    
#                     if contract:
#                         # Use payroll_calculation if contract exists
#                         payslip_data = payroll_calculation(
#                             employee=employee,
#                             start_date=start_date,
#                             end_date=end_date,
#                             wage=contract.wage
#                         )
#                     else:
#                         # Use calculate_payslip_from_attendance if no contract
#                         wage = instance.contract_wage or 0
#                         pay_head_data = instance.pay_head_data or {}
#                         if isinstance(pay_head_data, str):
#                             try:
#                                 pay_head_data = json.loads(pay_head_data)
#                             except:
#                                 pay_head_data = {}
                        
#                         payslip_data = calculate_payslip_from_attendance(
#                             employee=employee,
#                             start_date=start_date,
#                             end_date=end_date,
#                             wage=wage,
#                             housing_allowance=pay_head_data.get("housing_allowance", 0),
#                             transport_allowance=pay_head_data.get("transport_allowance", 0),
#                             other_allowance=pay_head_data.get("other_allowance", 0)
#                         )
                    
#                     # Update existing payslip with recalculated data
#                     instance.contract_wage = round(payslip_data.get('contract_wage', instance.contract_wage or 0), 2)
#                     instance.basic_pay = round(payslip_data.get('basic_pay', 0), 2)
#                     instance.gross_pay = round(payslip_data.get('gross_pay', 0), 2)
#                     instance.net_pay = round(payslip_data.get('net_pay', 0), 2)
#                     instance.deduction = round(payslip_data.get('total_deductions', 0), 2)
#                     instance.pay_head_data = payslip_data.get('pay_head_data', {})
#                     instance.save()
                    
#                     return HttpResponse(
#                         f"<script>window.location.href='/payroll/view-payslip/{instance.id}'</script>"
#                     )
#                 except Exception as e:
#                     print(f"[ADD_BONUS] Error recalculating payslip: {e}")
#                     import traceback
#                     traceback.print_exc()
#                     messages.warning(request, _("Payslip could not be recalculated: {}").format(str(e)))
#             return HttpResponse("<script>window.location.reload()</script>")
#     return render(
#         request,
#         "payroll/bonus/form.html",
#         {"form": form, "employee_id": employee_id, "payslip_id": payslip_id},
#     )

@login_required
@permission_required("payroll.add_allowance")
def add_bonus(request):
    employee_id = request.GET["employee_id"]
    payslip_id = request.GET.get("payslip_id")
    instance = None
    if payslip_id != "None" and payslip_id:
        try:
            instance = Payslip.objects.get(id=payslip_id)
        except Payslip.DoesNotExist:
            messages.error(request, _("Payslip not found."))
            return HttpResponse("<script>window.location.reload()</script>")
        form = forms.PayslipAllowanceForm(
            initial={"employee_id": employee_id, "date": instance.start_date}
        )
    else:
        form = forms.BonusForm(initial={"employee_id": employee_id})
    if request.method == "POST":
        form = forms.BonusForm(request.POST, initial={"employee_id": employee_id})
        contract = Contract.objects.filter(
            employee_id=employee_id, contract_status="active"
        ).first()
        employee = Employee.objects.filter(id=employee_id).first()
        if form.is_valid():
            form.save()
            messages.success(request, _("Bonus Added"))
            if payslip_id != "None" and payslip_id and instance:
                # CRITICAL: Check if this is an imported payslip - preserve Excel data
                pay_head_data = instance.pay_head_data or {}
                if isinstance(pay_head_data, str):
                    try:
                        pay_head_data = json.loads(pay_head_data)
                    except (json.JSONDecodeError, TypeError):
                        pay_head_data = {}
                
                is_imported = pay_head_data.get('is_imported', False)
                
                if is_imported:
                    # For imported payslips, preserve ALL Excel data and just update gross/net pay
                    print(f"[ADD_BONUS] Payslip is imported - preserving Excel data, updating gross/net pay")
                    
                    employee = instance.employee_id
                    start_date = instance.start_date
                    end_date = instance.end_date
                    
                    # Preserve all Excel values
                    excel_basic_pay = pay_head_data.get('basic_pay', instance.basic_pay or 0)
                    excel_housing = pay_head_data.get('housing_allowance', 0)
                    excel_transport = pay_head_data.get('transport_allowance', 0)
                    excel_other = pay_head_data.get('other_allowance', 0)
                    excel_overtime = pay_head_data.get('overtime', 0)
                    excel_salary_advance = pay_head_data.get('salary_advance', 0)
                    excel_bonus = pay_head_data.get('bonus', 0)
                    excel_lop = pay_head_data.get('loss_of_pay', pay_head_data.get('lop', 0))
                    excel_loan_recovery = pay_head_data.get('salary_advance_loan_recovery', 0)
                    excel_deduction = pay_head_data.get('deduction', 0)
                    
                    # Get bonuses from database (including the newly added one)
                    bonus_allowances = Allowance.objects.filter(
                        specific_employees=employee,
                        only_show_under_employee=True
                    ).filter(
                        Q(one_time_date__isnull=True) | 
                        Q(one_time_date__gte=start_date, one_time_date__lte=end_date)
                    )
                    
                    # Calculate total bonuses from database
                    db_bonus_total = 0
                    for allowance in bonus_allowances:
                        if "bonus" in allowance.title.lower():
                            if allowance.is_fixed:
                                db_bonus_total += float(allowance.amount or 0)
                            else:
                                rate = float(allowance.rate or 0)
                                db_bonus_total += (excel_basic_pay * rate) / 100
                    
                    # Total bonus = Excel bonus + Database bonuses
                    total_bonus = excel_bonus + db_bonus_total
                    
                    # CRITICAL: Build allowances list with bonuses from database (for template display)
                    allowances_list = pay_head_data.get('allowances', [])
                    if not isinstance(allowances_list, list):
                        allowances_list = []
                    
                    # Add bonuses from database to allowances list (deduplicate by title)
                    existing_bonus_titles = {a.get('title', '').lower() for a in allowances_list if 'bonus' in a.get('title', '').lower()}
                    for allowance in bonus_allowances:
                        if "bonus" in allowance.title.lower():
                            bonus_title = allowance.title
                            if bonus_title.lower() not in existing_bonus_titles:
                                amount = 0
                                if allowance.is_fixed:
                                    amount = float(allowance.amount or 0)
                                else:
                                    rate = float(allowance.rate or 0)
                                    amount = (excel_basic_pay * rate) / 100
                                
                                if amount > 0:
                                    allowances_list.append({
                                        "title": bonus_title,
                                        "amount": round(amount, 2),
                                        "id": allowance.id
                                    })
                                    existing_bonus_titles.add(bonus_title.lower())
                    
                    # Recalculate gross pay: Basic + Allowances + Bonuses
                    gross_pay = round(excel_basic_pay + excel_housing + excel_transport + excel_other + db_bonus_total, 2)
                    
                    # Recalculate net pay using Excel formula: (Gross Pay + Overtime + Salary Advance + Bonus) - (LOP + Loan Recovery + Deduction)
                    net_pay = round(
                        (gross_pay + excel_overtime + excel_salary_advance + total_bonus) - 
                        (excel_lop + excel_loan_recovery + excel_deduction),
                        2
                    )
                    
                    # Update pay_head_data
                    pay_head_data['gross_pay'] = gross_pay
                    pay_head_data['net_pay'] = net_pay
                    pay_head_data['bonus'] = total_bonus
                    pay_head_data['allowances'] = allowances_list  # CRITICAL: Update allowances list so bonuses appear in template
                    
                    # Update payslip model
                    instance.gross_pay = gross_pay
                    instance.net_pay = net_pay
                    instance.pay_head_data = pay_head_data
                    instance.save()
                    
                    print(f"[ADD_BONUS] Preserved Excel data - Gross Pay: {gross_pay}, Net Pay: {net_pay}, Total Bonus: {total_bonus}")
                    
                    return HttpResponse(
                        f"<script>window.location.href='/payroll/view-payslip/{instance.id}'</script>"
                    )
                else:
                    # For non-imported payslips, recalculate normally
                    try:
                        employee = instance.employee_id
                        start_date = instance.start_date
                        end_date = instance.end_date
                        
                        # Try to recalculate using payroll_calculation or calculate_payslip_from_attendance
                        contract = Contract.objects.filter(
                            employee_id=employee.id, contract_status="active"
                        ).first()
                        
                        if contract:
                            # Use payroll_calculation if contract exists
                            payslip_data = payroll_calculation(
                                employee=employee,
                                start_date=start_date,
                                end_date=end_date,
                                wage=contract.wage
                            )
                        else:
                            # Use calculate_payslip_from_attendance if no contract
                            wage = instance.contract_wage or 0
                            pay_head_data = instance.pay_head_data or {}
                            if isinstance(pay_head_data, str):
                                try:
                                    pay_head_data = json.loads(pay_head_data)
                                except:
                                    pay_head_data = {}
                            
                            payslip_data = calculate_payslip_from_attendance(
                                employee=employee,
                                start_date=start_date,
                                end_date=end_date,
                                wage=wage,
                                housing_allowance=pay_head_data.get("housing_allowance", 0),
                                transport_allowance=pay_head_data.get("transport_allowance", 0),
                                other_allowance=pay_head_data.get("other_allowance", 0)
                            )
                        
                        # Update existing payslip with recalculated data (same pattern as add_deduction)
                        instance.contract_wage = round(payslip_data.get('contract_wage', instance.contract_wage or 0), 2)
                        instance.basic_pay = round(payslip_data.get('basic_pay', 0), 2)
                        instance.gross_pay = round(payslip_data.get('gross_pay', 0), 2)
                        instance.net_pay = round(payslip_data.get('net_pay', 0), 2)
                        instance.deduction = round(payslip_data.get('total_deductions', 0), 2)
                        
                        # CRITICAL: Extract bonus from allowances and store in pay_head_data for table view
                        # This ensures bonus is visible in payslip_table.html (same concept as add_deduction extracts deductions)
                        pay_head_data = payslip_data.get('pay_head_data', {})
                        allowances_list = pay_head_data.get('allowances', [])
                        if not isinstance(allowances_list, list):
                            allowances_list = []
                        
                        # Extract bonus total from allowances list (for table view display)
                        bonus_total = 0
                        for allowance in allowances_list:
                            if allowance and isinstance(allowance, dict):
                                title = allowance.get('title', '').lower()
                                amount = float(allowance.get('amount', 0) or 0)
                                if 'bonus' in title:
                                    bonus_total += amount
                        
                        # Store bonus in pay_head_data for table view (payslip_table.html looks for pay_head_data.bonus)
                        pay_head_data['bonus'] = round(bonus_total, 2)
                        pay_head_data['allowances'] = allowances_list  # Preserve allowances list
                        
                        instance.pay_head_data = pay_head_data
                        instance.save()
                        
                        return HttpResponse(
                            f"<script>window.location.href='/payroll/view-payslip/{instance.id}'</script>"
                        )
                    except Exception as e:
                        print(f"[ADD_BONUS] Error recalculating payslip: {e}")
                        import traceback
                        traceback.print_exc()
                        messages.warning(request, _("Payslip could not be recalculated: {}").format(str(e)))
            return HttpResponse("<script>window.location.reload()</script>")
    return render(
        request,
        "payroll/bonus/form.html",
        {"form": form, "employee_id": employee_id, "payslip_id": payslip_id},
    )


@login_required
@permission_required("payroll.add_deduction")
def add_deduction(request):
    employee_id = request.GET["employee_id"]
    payslip_id = request.GET.get("payslip_id")
    
    # Safely get payslip instance
    try:
        instance = Payslip.objects.get(id=payslip_id)
    except Payslip.DoesNotExist:
        messages.error(request, _("Payslip not found."))
        return HttpResponse("<script>window.location.reload()</script>")

    if request.method == "POST":
        form = forms.PayslipDeductionForm(
            request.POST,
            initial={"employee_id": employee_id, "one_time_date": instance.start_date},
        )
        if form.is_valid():
            # Save the form to create the Deduction instance
            deduction_instance = form.save(commit=False)
            deduction_instance.only_show_under_employee = True
            deduction_instance.save()

            # Now that the Deduction instance is saved, add the related employees
            deduction_instance.specific_employees.set([employee_id])
            deduction_instance.include_active_employees = False
            deduction_instance.save()

            # Get employee and dates first (needed for both imported and non-imported payslips)
            employee = instance.employee_id
            start_date = instance.start_date
            end_date = instance.end_date
            
            # CRITICAL: Check if this is an imported payslip - preserve Excel data
            pay_head_data = instance.pay_head_data or {}
            if isinstance(pay_head_data, str):
                try:
                    import json
                    pay_head_data = json.loads(pay_head_data)
                except (json.JSONDecodeError, TypeError):
                    pay_head_data = {}
            
            is_imported = pay_head_data.get('is_imported', False)
            
            if is_imported:
                # For imported payslips, preserve ALL Excel data and just update net pay
                print(f"[ADD_DEDUCTION] Payslip is imported - preserving Excel data, only updating net pay")
                
                # Preserve all Excel values
                excel_basic_pay = pay_head_data.get('basic_pay', instance.basic_pay or 0)
                excel_gross_pay = pay_head_data.get('gross_pay', instance.gross_pay or 0)
                excel_overtime = pay_head_data.get('overtime', 0)
                excel_salary_advance = pay_head_data.get('salary_advance', 0)
                excel_bonus = pay_head_data.get('bonus', 0)
                excel_lop = pay_head_data.get('loss_of_pay', pay_head_data.get('lop', 0))
                excel_loan_recovery = pay_head_data.get('salary_advance_loan_recovery', 0)
                excel_deduction = pay_head_data.get('deduction', 0)
                
                # Get deductions from database (including the newly added one)
                deductions_queryset = Deduction.objects.filter(
                    specific_employees=employee,
                    only_show_under_employee=True
                )
                if start_date and end_date:
                    deductions_queryset = deductions_queryset.filter(
                        Q(one_time_date__isnull=True) | 
                        Q(one_time_date__gte=start_date, one_time_date__lte=end_date)
                    )
                
                # Calculate total deductions from database (excluding update_compensation deductions)
                db_deduction_total = 0
                for deduction in deductions_queryset:
                    if not deduction.update_compensation:  # Only count non-compensation deductions
                        if deduction.is_fixed:
                            db_deduction_total += float(deduction.amount or 0)
                        else:
                            # Calculate based on rate
                            base_amount = excel_basic_pay if deduction.based_on == "basic_pay" else excel_gross_pay
                            rate = float(deduction.rate or 0)
                            db_deduction_total += (base_amount * rate) / 100
                
                # Total deductions = Excel deductions (LOP + Loan Recovery + Deduction) + Database deductions
                total_deductions = round(excel_lop + excel_loan_recovery + excel_deduction + db_deduction_total, 2)
                
                # Recalculate net pay using Excel formula: (Gross Pay + Overtime + Salary Advance + Bonus) - (LOP + Loan Recovery + Deduction + DB Deductions)
                net_pay = round(
                    (excel_gross_pay + excel_overtime + excel_salary_advance + excel_bonus) - 
                    (excel_lop + excel_loan_recovery + excel_deduction + db_deduction_total),
                    2
                )
                
                # Update pay_head_data with new net pay and total deductions
                pay_head_data['net_pay'] = net_pay
                pay_head_data['total_deductions'] = total_deductions
                
                # Update payslip model
                instance.net_pay = net_pay
                instance.deduction = total_deductions
                instance.pay_head_data = pay_head_data
                instance.save()
                
                print(f"[ADD_DEDUCTION] Preserved Excel data - Net Pay: {net_pay}, Total Deductions: {total_deductions}")
                
                return HttpResponse(
                    f"<script>window.location.href='/payroll/view-payslip/{instance.id}'</script>"
                )
            else:
                # For non-imported payslips, recalculate normally
                try:
                    employee = instance.employee_id
                    start_date = instance.start_date
                    end_date = instance.end_date
                    
                    # Try to recalculate using payroll_calculation or calculate_payslip_from_attendance
                    contract = Contract.objects.filter(
                        employee_id=employee.id, contract_status="active"
                    ).first()
                    
                    if contract:
                        # Use payroll_calculation if contract exists
                        payslip_data = payroll_calculation(
                            employee=employee,
                            start_date=start_date,
                            end_date=end_date,
                            wage=contract.wage
                        )
                    else:
                        # Use calculate_payslip_from_attendance if no contract
                        wage = instance.contract_wage or 0
                        pay_head_data = instance.pay_head_data or {}
                        if isinstance(pay_head_data, str):
                            try:
                                pay_head_data = json.loads(pay_head_data)
                            except:
                                pay_head_data = {}
                        
                        payslip_data = calculate_payslip_from_attendance(
                            employee=employee,
                            start_date=start_date,
                            end_date=end_date,
                            wage=wage,
                            housing_allowance=pay_head_data.get("housing_allowance", 0),
                            transport_allowance=pay_head_data.get("transport_allowance", 0),
                            other_allowance=pay_head_data.get("other_allowance", 0)
                        )
                    
                    # Update existing payslip with recalculated data
                    instance.contract_wage = round(payslip_data.get('contract_wage', instance.contract_wage or 0), 2)
                    instance.basic_pay = round(payslip_data.get('basic_pay', 0), 2)
                    instance.gross_pay = round(payslip_data.get('gross_pay', 0), 2)
                    instance.net_pay = round(payslip_data.get('net_pay', 0), 2)
                    instance.deduction = round(payslip_data.get('total_deductions', 0), 2)
                    instance.pay_head_data = payslip_data.get('pay_head_data', {})
                    instance.save()
                    
                    return HttpResponse(
                        f"<script>window.location.href='/payroll/view-payslip/{instance.id}'</script>"
                    )
                except Exception as e:
                    print(f"[ADD_DEDUCTION] Error recalculating payslip: {e}")
                    import traceback
                    traceback.print_exc()
                    messages.warning(request, _("Payslip could not be recalculated: {}").format(str(e)))
                    return HttpResponse("<script>window.location.reload()</script>")

    else:
        form = forms.PayslipDeductionForm(
            initial={"employee_id": employee_id, "one_time_date": instance.start_date}
        )

    return render(
        request,
        "payroll/deduction/payslip_deduct.html",
        {"form": form, "employee_id": employee_id, "payslip_id": payslip_id},
    )



@login_required
@permission_required("payroll.view_loanaccount")
def view_loans(request):
    """
    This method is used to render template to disply all the loan records
    """
    records = LoanAccount.objects.all()
    loan = records.filter(type="loan")
    adv_salary = records.filter(type="advanced_salary")
    fine = records.filter(type="fine")

    fine_ids = json.dumps(list(fine.values_list("id", flat=True)))
    loan_ids = json.dumps(list(loan.values_list("id", flat=True)))
    adv_salary_ids = json.dumps(list(adv_salary.values_list("id", flat=True)))
    loan = sortby(request, loan, "sortby")
    adv_salary = sortby(request, adv_salary, "sortby")
    fine = sortby(request, fine, "sortby")
    filter_instance = LoanAccountFilter()
    return render(
        request,
        "payroll/loan/view_loan.html",
        {
            "records": paginator_qry(records, request.GET.get("page")),
            "loan": paginator_qry(loan, request.GET.get("lpage")),
            "adv_salary": paginator_qry(adv_salary, request.GET.get("apage")),
            "fine_ids": fine_ids,
            "loan_ids": loan_ids,
            "adv_salary_ids": adv_salary_ids,
            "fine": paginator_qry(fine, request.GET.get("fpage")),
            "f": filter_instance,
        },
    )


@login_required
@hx_request_required
def create_loan(request):
    """
    This method is used to create and update the loan instance
    """
    instance_id = eval_validate(str(request.GET.get("instance_id")))
    instance = LoanAccount.objects.filter(id=instance_id).first()
    form = forms.LoanAccountForm(instance=instance)
    if request.method == "POST":
        form = forms.LoanAccountForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Loan created/updated")
            return HttpResponse("<script>window.location.reload()</script>")
    return render(
        request, "payroll/loan/form.html", {"form": form, "instance_id": instance_id}
    )


@login_required
@permission_required("payroll.view_loanaccount")
def view_installments(request):
    """
    View install ments
    """
    loan_id = request.GET["loan_id"]
    loan = LoanAccount.objects.get(id=loan_id)
    installments = loan.deduction_ids.all()

    requests_ids_json = request.GET.get("instances_ids")
    if requests_ids_json:
        requests_ids = json.loads(requests_ids_json)
        previous_id, next_id = closest_numbers(requests_ids, int(loan_id))
    return render(
        request,
        "payroll/loan/installments.html",
        {
            "installments": installments,
            "loan": loan,
            "instances_ids": requests_ids_json,
            "previous": previous_id,
            "next": next_id,
        },
    )


@login_required
@permission_required("payroll.delete_loanaccount")
def delete_loan(request):
    """
    Delete loan
    """
    ids = request.GET.getlist("ids")
    loans = LoanAccount.objects.filter(id__in=ids)
    # This 👇 would'nt trigger the delete method in the model
    # loans.delete()
    for loan in loans:
        if (
            not loan.settled
            and not Payslip.objects.filter(
                installment_ids__in=list(
                    loan.deduction_ids.values_list("id", flat=True)
                )
            ).exists()
        ):
            loan.delete()
            messages.success(request, "Loan account deleted")
        else:
            messages.error(request, "Loan account cannot be deleted")
    return redirect(view_loans)


@login_required
@permission_required("payroll.view_loanaccount")
def edit_installment_amount(request):
    loan_id = request.GET.get("loan_id")
    ded_id = request.GET.get("ded_id")
    value = float(request.POST.get("amount")) if request.POST.get("amount") else 0

    loan = LoanAccount.objects.filter(id=loan_id).first()
    deductions = loan.deduction_ids.all().order_by("one_time_date")
    deduction = deductions.filter(id=ded_id).first()
    deductions_before = deductions.filter(one_time_date__lt=deduction.one_time_date)
    deductions_after = deductions.filter(one_time_date__gt=deduction.one_time_date)
    total_sum = deductions_before.aggregate(Sum("amount"))["amount__sum"] or 0

    balance_instalment = len(deductions_after) if len(deductions_after) != 0 else 1

    new_installment = (loan.loan_amount - total_sum - value) / balance_instalment
    new_installment = round(new_installment, 2)
    if total_sum + value > loan.loan_amount:
        value = loan.loan_amount - total_sum
        new_installment = 0

    if not deduction.installment_payslip():
        deduction.amount = value
        deduction.save()

        for item in deductions.filter(one_time_date__gt=deduction.one_time_date):
            item.amount = new_installment
            item.save()

        if len(deductions_after) == 0 and new_installment != 0:
            date = get_next_month_same_date(deduction.one_time_date)
            installment = create_deductions(loan, new_installment, date)
            loan.deduction_ids.add(installment)

        messages.success(request, "Installment amount updated successfully")
    else:
        messages.error(request, "Cannot change paid installments ")

    return render(
        request,
        "payroll/loan/installments.html",
        {
            "installments": loan.deduction_ids.all(),
            "loan": loan,
        },
    )


@login_required
@hx_request_required
@permission_required("payroll.view_loanaccount")
def search_loan(request):
    """
    Search loan method
    """
    records = LoanAccountFilter(request.GET).qs
    loan = records.filter(type="loan")
    adv_salary = records.filter(type="advanced_salary")
    fine = records.filter(type="fine")

    fine_ids = json.dumps(list(fine.values_list("id", flat=True)))
    loan_ids = json.dumps(list(loan.values_list("id", flat=True)))
    adv_salary_ids = json.dumps(list(adv_salary.values_list("id", flat=True)))
    loan = sortby(request, loan, "sortby")
    adv_salary = sortby(request, adv_salary, "sortby")
    fine = sortby(request, fine, "sortby")

    data_dict = parse_qs(request.GET.urlencode())
    get_key_instances(LoanAccount, data_dict)
    view = request.GET.get("view")
    template = "payroll/loan/records_card.html"
    if view == "list":
        template = "payroll/loan/records_list.html"
    return render(
        request,
        template,
        {
            "records": paginator_qry(records, request.GET.get("page")),
            "loan": paginator_qry(loan, request.GET.get("lpage")),
            "adv_salary": paginator_qry(adv_salary, request.GET.get("apage")),
            "fine": paginator_qry(fine, request.GET.get("fpage")),
            "fine_ids": fine_ids,
            "loan_ids": loan_ids,
            "adv_salary_ids": adv_salary_ids,
            "filter_dict": data_dict,
            "pd": request.GET.urlencode(),
        },
    )


@login_required
@permission_required("payroll.add_loanaccount")
def asset_fine(request):
    """
    Add asset fine method
    """
    if apps.is_installed("asset"):
        Asset = get_horilla_model_class(app_label="asset", model="asset")
    asset_id = request.GET["asset_id"]
    employee_id = request.GET["employee_id"]
    asset = Asset.objects.get(id=asset_id)
    employee = Employee.objects.get(id=employee_id)
    form = forms.AssetFineForm()
    if request.method == "POST":
        form = forms.AssetFineForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.employee_id = employee
            instance.type = "fine"
            instance.provided_date = date.today()
            instance.asset_id = asset
            instance.save()
            messages.success(request, _("Asset fine added"))
            return HttpResponse(
                "<script>$('#dynamicCreateModal').toggleClass('oh-modal--show'); $('#reloadMessagesButton').click();</script>"
            )  # 880
    return render(
        request,
        "payroll/asset_fine/form.html",
        {"form": form, "asset_id": asset_id, "employee_id": employee_id},
    )


@login_required
def view_reimbursement(request):
    """
    This method is used to render template to view reimbursements
    """
    reimbursement_exists = False
    if Reimbursement.objects.exists():
        reimbursement_exists = True
    if request.GET:
        filter_object = ReimbursementFilter(request.GET)
    else:
        filter_object = ReimbursementFilter({"status": "requested"})
    requests = filter_own_records(
        request, filter_object.qs, "payroll.view_reimbursement"
    )
    reimbursements = requests.filter(type="reimbursement")
    leave_encashments = requests.filter(type="leave_encashment")
    bonus_encashment = requests.filter(type="bonus_encashment")
    data_dict = {"status": ["requested"]}
    view = request.GET.get("view")
    template = "payroll/reimbursement/view_reimbursement.html"

    return render(
        request,
        template,
        {
            "requests": paginator_qry(requests, request.GET.get("page")),
            "reimbursements": paginator_qry(reimbursements, request.GET.get("rpage")),
            "leave_encashments": paginator_qry(
                leave_encashments, request.GET.get("lpage")
            ),
            "bonus_encashments": paginator_qry(
                bonus_encashment, request.GET.get("bpage")
            ),
            "f": filter_object,
            "pd": request.GET.urlencode(),
            "filter_dict": data_dict,
            "view": view,
            "reimbursement_exists": reimbursement_exists,
        },
    )


@login_required
@hx_request_required
def create_reimbursement(request):
    """
    Create or update a reimbursement entry.
    """
    instance = None
    instance_id = request.GET.get("instance_id")

    if instance_id:
        instance = Reimbursement.objects.filter(id=instance_id).first()

    if request.method == "POST":
        form = forms.ReimbursementForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Reimbursement saved successfully")
            return HttpResponse(status=204, headers={"HX-Refresh": "true"})
    else:
        form = forms.ReimbursementForm(instance=instance)

    return render(request, "payroll/reimbursement/form.html", {"form": form})


@login_required
@hx_request_required
def search_reimbursement(request):
    """
    This method is used to search/filter reimbursement
    """
    requests = ReimbursementFilter(request.GET).qs
    requests = filter_own_records(request, requests, "payroll.view_reimbursement")
    data_dict = parse_qs(request.GET.urlencode())
    reimbursements = requests.filter(type="reimbursement")
    leave_encashments = requests.filter(type="leave_encashment")
    bonus_encashment = requests.filter(type="bonus_encashment")
    reimbursements_ids = json.dumps(list(reimbursements.values_list("id", flat=True)))
    leave_encashments_ids = json.dumps(
        list(leave_encashments.values_list("id", flat=True))
    )
    bonus_encashment_ids = json.dumps(
        list(bonus_encashment.values_list("id", flat=True))
    )
    reimbursements = sortby(request, reimbursements, "sortby")
    leave_encashments = sortby(request, leave_encashments, "sortby")
    bonus_encashment = sortby(request, bonus_encashment, "sortby")
    view = request.GET.get("view")
    template = "payroll/reimbursement/request_cards.html"
    if view == "list":
        template = "payroll/reimbursement/reimbursement_list.html"
    get_key_instances(Reimbursement, data_dict)

    return render(
        request,
        template,
        {
            "requests": paginator_qry(requests, request.GET.get("page")),
            "reimbursements": paginator_qry(reimbursements, request.GET.get("rpage")),
            "leave_encashments": paginator_qry(
                leave_encashments, request.GET.get("lpage")
            ),
            "bonus_encashments": paginator_qry(
                bonus_encashment, request.GET.get("bpage")
            ),
            "filter_dict": data_dict,
            "pd": request.GET.urlencode(),
            "reimbursements_ids": reimbursements_ids,
            "leave_encashments_ids": leave_encashments_ids,
            "bonus_encashment_ids": bonus_encashment_ids,
        },
    )


@login_required
def get_assigned_leaves(request):
    """
    This method is used to return assigned leaves of the employee
    in Json
    """
    if apps.is_installed("leave"):
        AvailableLeave = get_horilla_model_class(
            app_label="leave", model="availableleave"
        )

    assigned_leaves = (
        AvailableLeave.objects.filter(
            employee_id__id=request.GET["employeeId"],
            total_leave_days__gte=1,
            leave_type_id__is_encashable=True,
        )
        .values(
            "leave_type_id__name",
            "available_days",
            "carryforward_days",
            "leave_type_id__id",
        )
        .distinct()
    )
    return JsonResponse(list(assigned_leaves), safe=False)


@login_required
@permission_required("payroll.change_reimbursement")
def approve_reimbursements(request):
    """
    This method is used to approve or reject the reimbursement request
    """
    ids = request.GET.getlist("ids")
    status = request.GET["status"]
    if status == "canceled":
        status = "rejected"
    amount = (
        eval_validate(request.GET.get("amount")) if request.GET.get("amount") else 0
    )
    amount = max(0, amount)
    reimbursements = Reimbursement.objects.filter(id__in=ids)
    if status and len(status):
        for reimbursement in reimbursements:
            if reimbursement.type == "leave_encashment":
                reimbursement.amount = amount
            elif reimbursement.type == "bonus_encashment":
                reimbursement.amount = amount

            emp = reimbursement.employee_id
            reimbursement.status = status
            reimbursement.save()
            if reimbursement.status == "requested":
                if not (messages.get_messages(request)._queued_messages):
                    messages.info(request, _("Please check the data you provided."))
            else:
                messages.success(
                    request,
                    _(f"Request {reimbursement.get_status_display()} successfully"),
                )
        if status == "rejected":
            notify.send(
                request.user.employee_get,
                recipient=emp.employee_user_id,
                verb="Your reimbursement request has been rejected.",
                verb_ar="تم رفض طلب استرداد النفقات الخاص بك.",
                verb_de="Ihr Erstattungsantrag wurde abgelehnt.",
                verb_es="Su solicitud de reembolso ha sido rechazada.",
                verb_fr="Votre demande de remboursement a été rejetée.",
                redirect=reverse("view-reimbursement") + f"?id={reimbursement.id}",
                icon="checkmark",
            )
        else:
            notify.send(
                request.user.employee_get,
                recipient=emp.employee_user_id,
                verb="Your reimbursement request has been approved.",
                verb_ar="تمت الموافقة على طلب استرداد نفقاتك.",
                verb_de="Ihr Rückerstattungsantrag wurde genehmigt.",
                verb_es="Se ha aprobado tu solicitud de reembolso.",
                verb_fr="Votre demande de remboursement a été approuvée.",
                redirect=reverse("view-reimbursement") + f"?id={reimbursement.id}",
                icon="checkmark",
            )
    return redirect(view_reimbursement)


@login_required
@permission_required("payroll.delete_reimbursement")
def delete_reimbursements(request):
    """
    This method is used to delete the reimbursements
    """
    ids = request.GET.getlist("ids")
    reimbursements = Reimbursement.objects.filter(id__in=ids)
    for reimbursement in reimbursements:
        user = reimbursement.employee_id.employee_user_id
    reimbursements.delete()
    messages.success(request, "Reimbursements deleted")
    notify.send(
        request.user.employee_get,
        recipient=user,
        verb="Your reimbursement request has been deleted.",
        verb_ar="تم حذف طلب استرداد نفقاتك.",
        verb_de="Ihr Rückerstattungsantrag wurde gelöscht.",
        verb_es="Tu solicitud de reembolso ha sido eliminada.",
        verb_fr="Votre demande de remboursement a été supprimée.",
        redirect="/",
        icon="trash",
    )

    return redirect(view_reimbursement)


@login_required
@owner_can_enter("payroll.view_reimbursement", Reimbursement, True)
def reimbursement_individual_view(request, instance_id):
    """
    This method is used to render the individual view of reimbursement object
    """
    reimbursement = Reimbursement.objects.get(id=instance_id)
    requests_ids_json = request.GET.get("instances_ids")
    if requests_ids_json:
        requests_ids = json.loads(requests_ids_json)
        previous_id, next_id = closest_numbers(requests_ids, instance_id)
    context = {
        "reimbursement": reimbursement,
        "instances_ids": requests_ids_json,
        "previous": previous_id,
        "next": next_id,
    }
    return render(
        request,
        "payroll/reimbursement/reimbursenent_individual.html",
        context,
    )


@login_required
@owner_can_enter("payroll.view_reimbursement", Reimbursement, True)
def reimbursement_attachments(request, instance_id):
    """
    This method is used to render all the attachements under the reimbursement object
    """
    reimbursement = Reimbursement.objects.get(id=instance_id)
    return render(
        request,
        "payroll/reimbursement/attachments.html",
        {"reimbursement": reimbursement},
    )


@login_required
@owner_can_enter("payroll.delete_reimbursement", Reimbursement, True)
def delete_attachments(request, _reimbursement_id):
    """
    This mehtod is used to delete the attachements
    """
    ids = request.GET.getlist("ids")
    ReimbursementMultipleAttachment.objects.filter(id__in=ids).delete()
    messages.success(request, "Attachment deleted")
    return redirect(view_reimbursement)


@login_required
@permission_required("payroll.view_payslip")
def get_contribution_report(request):
    """
    This method is used to get the contribution report
    """
    employee_id = request.GET.get("employee_id")
    contribution_deductions = []
    if employee_id:
        pay_heads = Payslip.objects.filter(employee_id__id=employee_id).values_list(
            "pay_head_data", flat=True
        )
        deductions = []
        for head in pay_heads:
            for deduction in head["gross_pay_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["basic_pay_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["pretax_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["post_tax_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["tax_deductions"]:
                if deduction.get("deduction_id"):
                    deductions.append(deduction)
            for deduction in head["net_deductions"]:
                deductions.append(deduction)

        deductions.sort(key=lambda x: x["deduction_id"])
        grouped_deductions = {
            key: list(group)
            for key, group in groupby(deductions, key=lambda x: x["deduction_id"])
        }

        for deduction_id, group in grouped_deductions.items():
            title = group[0]["title"]
            employee_contribution = sum(item.get("amount", 0) for item in group)
            employer_contribution = sum(
                item.get("employer_contribution_amount", 0) for item in group
            )
            total_contribution = employee_contribution + employer_contribution
            if employer_contribution > 0:
                contribution_deductions.append(
                    {
                        "deduction_id": deduction_id,
                        "title": title,
                        "employee_contribution": employee_contribution,
                        "employer_contribution": employer_contribution,
                        "total_contribution": total_contribution,
                    }
                )
    return render(
        request,
        "payroll/dashboard/contribution.html",
        {"contribution_deductions": contribution_deductions},
    )


def all_deductions(pay_head):

    extracted_items = []

    potential_lists = [
        "basic_pay_deductions",
        "gross_pay_deductions",
        "pretax_deductions",
        "post_tax_deductions",
        "tax_deductions",
        "net_deductions",
    ]

    for list_name in potential_lists:
        if list_name in pay_head.keys():
            for item in pay_head[list_name]:
                if "deduction_id" in item:
                    extracted_items.append(item)

    return extracted_items


@login_required
def payslip_detailed_export_data(request):
    """
    This view create the data for exporting payslip data based on selected fields and filters,
    """
    choices_mapping = {
        "draft": _("Draft"),
        "review_ongoing": _("Review Ongoing"),
        "confirmed": _("Confirmed"),
        "paid": _("Paid"),
    }
    selected_columns = []
    payslips_data = []
    totals = {}
    payslips = PayslipFilter(request.GET).qs
    selected_fields = request.GET.getlist("selected_fields")
    form = forms.PayslipExportColumnForm()

    allowances = Allowance.objects.all()
    deductions = Deduction.objects.all()

    if not selected_fields:
        selected_fields = form.fields["selected_fields"].initial

    for field in forms.excel_columns:
        value, key = field

        if value in selected_fields:
            selected_columns.append((value, key))

    selected_columns += [
        (value.title, value.title)
        for value in allowances.filter(
            one_time_date__isnull=True, include_active_employees=True
        )
    ]
    selected_columns += [
        ("other_allowances", "Other Allowances"),
        ("total_allowances", "Total Allowances"),
    ]

    selected_columns += [
        (value.title, value.title)
        for value in deductions.filter(
            one_time_date__isnull=True,
            include_active_employees=True,
            update_compensation__isnull=True,
        )
    ]
    selected_columns += [
        ("federal_tax", "Federal Tax"),
        ("other_deductions", "Other Deductions"),
        ("total_deductions", "Total Deductions"),
    ]

    allowance_totals = {
        column_name.title: 0
        for column_name in allowances.filter(
            one_time_date__isnull=True,
            include_active_employees=True,
        )
    }

    deduction_totals = {
        column_name.title: 0
        for column_name in deductions.filter(
            one_time_date__isnull=True,
            include_active_employees=True,
            update_compensation__isnull=True,
        )
    }

    other_totals = {
        "Other Allowances": 0,
        "Other Deductions": 0,
        "Total Allowances": 0,
        "Total Deductions": 0,
        "Net Pay": 0,
        "Gross Pay": 0,
        "Federal Tax": 0,
    }

    totals.update(allowance_totals)
    totals.update(deduction_totals)
    totals.update(other_totals)
    for payslip in payslips:
        payslip_data = {}
        other_allowances_sum = 0
        other_deductions_sum = 0
        total_allowance = 0
        total_deduction = 0
        total_federal_tax = 0

        federal_tax = payslip.pay_head_data["federal_tax"]
        total_federal_tax += federal_tax

        allos = payslip.pay_head_data["allowances"]
        deducts = all_deductions(payslip.pay_head_data)

        if allos:
            for allowance in allos:
                if not any(
                    str(allowance["title"]) == str(column_name)
                    for item, column_name in selected_columns
                ):
                    other_allowances_sum += (
                        allowance["amount"] if allowance["amount"] is not None else 0
                    )
                total_allowance += allowance["amount"]

        if deducts:
            for deduction in deducts:
                if not any(
                    str(deduction["title"]) == str(column_name)
                    for item, column_name in selected_columns
                ):
                    other_deductions_sum += (
                        deduction["amount"] if deduction["amount"] is not None else 0
                    )
                total_deduction += deduction["amount"]

        for column_value, column_name in selected_columns:
            nested_attributes = column_value.split("__")
            value = payslip
            for attr in nested_attributes:
                value = getattr(value, attr, None)
                if value is None:
                    break
            data = str(value) if value is not None else ""
            if column_name == "Status":
                data = choices_mapping.get(value, "")

            if isinstance(value, date):
                date_format = request.user.employee_get.get_date_format()
                start_date = datetime.strptime(str(value), "%Y-%m-%d").date()

                for format_name, format_string in HORILLA_DATE_FORMATS.items():
                    if format_name == date_format:
                        data = start_date.strftime(format_string)
            else:
                data = str(value) if value is not None else ""

            if allos:
                for allowance in allos:
                    if str(allowance["title"]) == str(column_name):
                        data = (
                            float(allowance["amount"])
                            if allowance["title"] is not None
                            else 0
                        )

            if deducts:
                for deduction in deducts:
                    if str(deduction["title"]) == str(column_name):
                        data = (
                            float(deduction["amount"])
                            if deduction["title"] is not None
                            else 0
                        )

            payslip_data[column_name] = data
            if column_name in totals:
                try:
                    totals[column_name] += float(data)
                except ValueError:
                    pass

        payslip_data["Other Allowances"] = other_allowances_sum
        payslip_data["Other Deductions"] = other_deductions_sum
        payslip_data["Total Allowances"] = total_allowance
        payslip_data["Total Deductions"] = total_deduction
        payslip_data["Federal Tax"] = federal_tax

        totals["Other Allowances"] += other_allowances_sum
        totals["Other Deductions"] += other_deductions_sum
        totals["Total Allowances"] += total_allowance
        totals["Total Deductions"] += total_deduction
        totals["Federal Tax"] += federal_tax

        payslips_data.append(payslip_data)

    totals_row = {}

    for item, column_name in selected_columns:
        if column_name in totals:
            totals_row[column_name] = totals[column_name]
        else:
            totals_row[column_name] = "-"

    totals_row["Other Allowances"] = totals["Other Allowances"]
    totals_row["Other Deductions"] = totals["Other Deductions"]
    totals_row["Total Allowances"] = totals["Total Allowances"]
    totals_row["Total Deductions"] = totals["Total Deductions"]
    totals_row["Employee"] = "Total"

    payslips_data.append(totals_row)

    return {
        "payslips_data": payslips_data,
        "selected_columns": selected_columns,
        "allowances": list(
            allowances.filter(
                one_time_date__isnull=True,
                include_active_employees=True,
            ).values_list("title", flat=True)
        ),
        "deductions": list(
            deductions.filter(
                one_time_date__isnull=True,
                include_active_employees=True,
                update_compensation__isnull=True,
            ).values_list("title", flat=True)
        ),
    }


@login_required
@permission_required("payroll.change_payslip")
def payslip_detailed_export(request):
    """
    Generate an Excel file for download containing detailed payslip data based on
    filters.

    Args:
        request (HttpRequest): The incoming HTTP request object.

    Returns:
        HttpResponse: A response object with the Excel file as an attachment.
    """

    if request.META.get("HTTP_HX_REQUEST"):
        return render(
            request,
            "payroll/payslip/payslip_export_filter.html",
            {
                "export_column": forms.PayslipExportColumnForm(),
                "export_filter": PayslipFilter(request.GET),
                "report": True,
            },
        )

    export_data = payslip_detailed_export_data(request)
    payslips_data = export_data["payslips_data"]
    selected_columns = export_data["selected_columns"]
    allowances = export_data["allowances"]
    deductions = export_data["deductions"]
    today_date = date.today().strftime("%Y-%m-%d")
    file_name = f"Payslip_excel_{today_date}.xlsx"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    right_border = Border(right=Side(style="thin"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Payslips"

    header_row = [col_name for _, col_name in selected_columns]
    allowances_header = allowances + ["Other Allowances", "Total Allowances"]
    deductions_header = deductions + [
        "Federal Tax",
        "Other Deductions",
        "Total Deductions",
    ]

    basic_cols = len(header_row) - len(allowances_header) - len(deductions_header)
    allowance_cols = len(allowances_header)
    deduction_cols = len(deductions_header)

    merged_sections = [
        (1, basic_cols, "Employee Details", "0000FF"),
        (basic_cols + 1, basic_cols + allowance_cols, "Allowances", "008000"),
        (
            basic_cols + allowance_cols + 1,
            basic_cols + allowance_cols + deduction_cols,
            "Deductions",
            "FF0000",
        ),
    ]

    bold_cols = [
        1,
        basic_cols + allowance_cols,
        basic_cols + allowance_cols + deduction_cols,
    ]

    for start_col, end_col, title, color in merged_sections:
        ws.merge_cells(
            start_row=1, start_column=start_col, end_row=1, end_column=end_col
        )
        cell = ws.cell(row=1, column=start_col, value=title)
        cell.font = Font(color=color, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

        if end_col <= len(header_row):
            ws.cell(row=1, column=end_col).border = thin_border + right_border
    last_row = len(payslips_data) + 2
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[last_row].height = 25

    subheaders = [
        (header_row[:basic_cols], Font(bold=True, color="0000FF")),
        (allowances_header, Font(bold=True, color="008000")),
        (deductions_header, Font(bold=True, color="FF0000")),
    ]

    col_num = 1
    for subheader, font in subheaders:
        for header in subheader:
            cell = ws.cell(row=2, column=col_num, value=str(header))
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            col_num += 1

    for row_num, payslip_data in enumerate(payslips_data, 3):
        for col_num, header in enumerate(header_row, 1):
            cell = ws.cell(
                row=row_num, column=col_num, value=payslip_data.get(header, "")
            )
            if row_num == last_row:
                cell.font = Font(bold=True, color="800080")
                cell.alignment = Alignment(horizontal="right")
            elif col_num in bold_cols:
                cell.font = Font(bold=True)
            cell.border = thin_border

    for col_num, _ in enumerate(header_row, 1):
        max_length = max(
            len(str(cell.value))
            for cell in ws[get_column_letter(col_num)]
            if cell.value is not None
        )
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    ws.freeze_panes = ws["B3"]

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={file_name}.xlsx"
    wb.save(response)

    return response




# Assuming these imports are correct based on the context
# from payroll.models import Payslip, Employee 
# from payroll.forms import PayslipImportForm 


@login_required
@permission_required("payroll.run_payroll")
def run_payroll(request):
    """
    Renders the main payroll running page with dynamic year and month status.
    Requires the user to be logged in and have the 'payroll.run_payroll' permission.
    """
    try:
        current_year = date.today().year
        selected_year = int(request.GET.get('year', current_year))
    except (ValueError, TypeError):
        selected_year = date.today().year

    months_data = []
    
    # Calculate start/end dates for all months in the selected year
    month_ranges = get_month_start_end(selected_year)
    
    for i, (start_date, end_date) in enumerate(month_ranges):
        month_num = i + 1
        month_name = calendar.month_name[month_num]
        
        # Check status: If ANY payslip exists for this period, mark as completed
        # You might want to refine this (e.g., only if *all* active employees have payslips, or specific status)
        # For now, simplistic check: if payslips exist, it's "completed" (or in progress), else "pending"
        payslips_exist = Payslip.objects.filter(
            start_date__gte=start_date, 
            end_date__lte=end_date
        ).exists()
        
        status = "completed" if payslips_exist else "pending"
        
        # Format date range for display (e.g., "Jan 1 - Jan 31")
        date_range = f"{start_date.strftime('%b %d')} - {end_date.strftime('%b %d')}"
        
        months_data.append({
            'name': f"{month_name} {selected_year}",
            'range': date_range,
            'status': status,
            'month_num': month_num # Useful for links if needed
        })

    context = {
        'months': months_data,
        'selected_year': selected_year,
    }
    return render(request, "payroll/run_payroll/run_payroll.html", context)


def calculate_payslip_from_attendance(employee, start_date, end_date, wage, housing_allowance=0, transport_allowance=0, other_allowance=0):
    """
    Calculate payslip for employee based on attendance/leave records without requiring contract.
    Similar to Keka's approach - calculates directly from WorkRecords and Attendance.
    
    Args:
        employee: Employee instance
        start_date: Start date of payroll period
        end_date: End date of payroll period
        wage: Basic wage/salary for the employee
        housing_allowance: Housing allowance amount
        transport_allowance: Transport allowance amount
        other_allowance: Other allowance amount
    
    Returns:
        dict: Calculated payslip data
    """
    import logging
    # Use print as requested instead of logger
    
    print(f"[PAYSLIP_CALC] Starting calculation for Employee: {employee.badge_id} ({employee.get_full_name()})")
    print(f"[PAYSLIP_CALC] Period: {start_date} to {end_date}")
    print(f"[PAYSLIP_CALC] Wage: {wage}, Allowances - Housing: {housing_allowance}, Transport: {transport_allowance}, Other: {other_allowance}")
    
    # Get all dates in the period
    date_range = get_date_range(start_date, end_date)
    print(f"[PAYSLIP_CALC] Total days in period: {len(date_range)}")
    
    # Get employee company for filtering holidays
    employee_company = None
    department_name = "-"
    if hasattr(employee, 'employee_work_info') and employee.employee_work_info:
        employee_company = getattr(employee.employee_work_info, 'company_id', None)
        if employee.employee_work_info.department_id:
             department_name = employee.employee_work_info.department_id.department
        print(f"[PAYSLIP_CALC] Employee Company: {employee_company}")
    
    # Get holidays in the period (filtered by company if available)
    holiday_query = Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
    if employee_company:
        holiday_query &= Q(company_id=employee_company)
    all_holidays = Holidays.objects.filter(holiday_query)
    print(f"[PAYSLIP_CALC] Found {all_holidays.count()} holidays in period")
    
    holiday_dates_set = set()
    for hol in all_holidays:
        hol_start = max(hol.start_date, start_date)
        hol_end = min(hol.end_date or hol.start_date, end_date)
        day = hol_start
        while day <= hol_end:
            holiday_dates_set.add(day)
            day += timedelta(days=1)
    print(f"[PAYSLIP_CALC] Holiday dates count: {len(holiday_dates_set)}")
    
    # Get work records for the period
    work_records = WorkRecords.objects.filter(
        date__in=date_range,
        employee_id=employee
    )
    print(f"[PAYSLIP_CALC] Found {work_records.count()} work records")
    wr_dict = {wr.date: wr for wr in work_records}
    
    # Also get attendance records for validation
    attendance_records = Attendance.objects.filter(
        employee_id=employee,
        attendance_date__range=(start_date, end_date),
        attendance_validated=True
    )
    print(f"[PAYSLIP_CALC] Found {attendance_records.count()} validated attendance records")
    attendance_dates = set(att.attendance_date for att in attendance_records)
    
    # Calculate working days, paid days, and leave days
    total_days = len(date_range)
    present_days = 0
    half_day_present = 0
    absent_days = 0
    leave_full_days = 0
    leave_half_days = 0
    weekly_off_days = 0
    holiday_days = len(holiday_dates_set)
    overtime_hours = 0
    
    # Calculate overtime from attendance
    for att in attendance_records:
        if att.overtime_second:
            overtime_hours += att.overtime_second / 3600
    
    # Get employee shift for working day calculation
    emp_shift = None
    if hasattr(employee, "employee_work_info") and employee.employee_work_info:
        emp_shift = getattr(employee.employee_work_info, "shift_id", None)
    print(f"[PAYSLIP_CALC] Employee Shift: {emp_shift}")
    
    # Get working days data for accurate calculation
    working_days_data = get_working_days(start_date, end_date, company=employee_company)
    total_working_days = working_days_data.get("total_working_days", 0)
    working_days_list = working_days_data.get("working_days_on", [])
    print(f"[PAYSLIP_CALC] Total working days: {total_working_days}")
    
    # Process each date in the period
    for date_obj in date_range:
        if date_obj in holiday_dates_set:
            # Check if holiday is overlapping with a leave request or work record?
            # Usually holidays are paid.
            continue
        
        wr = wr_dict.get(date_obj)
        
        if wr:
            if wr.work_record_type == "FDP":
                present_days += 1
            elif wr.work_record_type == "HDP":
                half_day_present += 0.5
            elif wr.work_record_type == "ABS":
                absent_days += 1
            elif wr.work_record_type == "LFD":
                leave_full_days += 1
            elif wr.work_record_type == "LHD":
                leave_half_days += 0.5
            elif wr.work_record_type == "WOF":
                weekly_off_days += 1
        else:
            # No work record - check if it's a working day
            is_working_day = date_obj in working_days_list
            
            if is_working_day:
                # Check if there's attendance for this day
                if date_obj in attendance_dates:
                    present_days += 1
                    # print(f"[PAYSLIP_CALC] Date {date_obj}: Present (attendance found)")
                else:
                    # Check for leaves (get_leaves logic used separately or incorporated here)
                    # Simplified assumption: if no attendance and no work record, it might be absent
                    # unless we check for leave requests explicitly here.
                     # Ideally we should use get_leaves() or similar if work records aren't fully populated.
                    absent_days += 1
                    # print(f"[PAYSLIP_CALC] Date {date_obj}: Absent (no attendance)")
            else:
                weekly_off_days += 1
                # print(f"[PAYSLIP_CALC] Date {date_obj}: Weekly off")
    
    print(f"[PAYSLIP_CALC] Attendance Summary - Present: {present_days}, Half-day: {half_day_present}, "
                f"Absent: {absent_days}, Leave Full: {leave_full_days}, Leave Half: {leave_half_days}, "
                f"Weekly Off: {weekly_off_days}, Holidays: {holiday_days}")
    
    # Calculate paid days (present + half days + holidays + weekly offs + paid leaves)
    # Note: Leave full days and half days are typically unpaid unless specified otherwise.
    # We will assume LFD/LHD in WorkRecord are UNPAID unless we check payment type involving complexity.
    # But usually 'paid' leave is treated as Present or a specific paid type.
    # Let's assume standard logic: WorkRecords capture the final status.
    
    paid_days = present_days + half_day_present + holiday_days + weekly_off_days
    # If there are paid leaves, they should be added.
    # Let's check if LFD is paid or unpaid. Usually LFD in WorkRecord implies taken leave. 
    # Whether it is paid depends on leave type. Here we assume unpaid for simplification or LOP.
    
    # Use user Requirement Fields:
    # LOP (Loss of Pay)Days = Absent Days + Unpaid Leave Days
    unpaid_days = absent_days + leave_full_days + (leave_half_days * 0.5)
    
    print(f"[PAYSLIP_CALC] Paid days: {paid_days}, Unpaid days: {unpaid_days}")
    
    # Calculate basic pay (based on paid days)
    # Use total working days for per-day calculation
    # Standard formula: (Basic / Total Days in Month) * Paid Days
    
    days_in_month = (end_date - start_date).days + 1
    # Or use 30 as standard? Let's use actual days in range.
    
    if days_in_month > 0:
        per_day_wage = wage / days_in_month
    else:
        per_day_wage = 0
        
    print(f"[PAYSLIP_CALC] Per day wage: {per_day_wage:.2f}")
    
    # Basic Pay for the period (Actual Earnings vs Contract Wage)
    # Actually, often Basic Pay refers to the Contract Basic. 
    # Earned Basic Pay = (Contract Basic / Total Days) * Paid Days
    earned_basic_pay = per_day_wage * paid_days
    
    loss_of_pay_amount = per_day_wage * unpaid_days
    
    print(f"[PAYSLIP_CALC] Earned Basic pay: {earned_basic_pay:.2f}, Loss of pay: {loss_of_pay_amount:.2f}")
    
    # Calculate allowances
    # Are allowances fixed or pro-rated? Usually fixed but let's assume fixed for now as per previous code.
    total_allowances = housing_allowance + transport_allowance + other_allowance
    
    # Gross Pay = Earned Basic + Allowances - LOP? Or Earned Basic + Allowances?
    # Usually: Gross Pay = Basic + Allowances
    # But if there is LOP, it reduces the total earnings.
    # Let's calculating: Gross Pay (Contract) vs Gross Pay (Earned)
    
    gross_pay = earned_basic_pay + total_allowances
    
    print(f"[PAYSLIP_CALC] Total allowances: {total_allowances:.2f}, Gross pay: {gross_pay:.2f}")
    
    # Deductions
    salary_advance = 0 # Placeholder, need DB fetch if exists
    salary_advance_loan_recovery = 0 # Placeholder
    bonus = 0
    other_deductions = 0
    
    total_deductions = salary_advance_loan_recovery + other_deductions 
    # Note: LOP is already deducted by reducing Basic Pay, 
    # OR we show Full Basic and then show LOP as deduction.
    # Requirement asks for "LOP" column. So maybe we should show Full Basic in one column
    # and LOP amount in another? 
    # "Basic Pay" usually means Earned Basic in payslip context.
    # Let's stick to: Net Pay = Gross Pay - Deductions.
    # If Gross Pay is calculated on Paid Days, LOP is implicit.
    # If we want explicit LOP deduction:
    # Gross (Full) = Wage + Allowances
    # Deduction = LOP Amount + Others
    # Net = Gross (Full) - Deduction.
    # Let's assume the user wants checkable LOP.
    
    # Standard Keka/Hr approach:
    # Earnings: Basic (Earned), HRA, ...
    # Deductions: PF, Tax, etc.
    # LOP is often shown as days.
    
    # Let's calculate proper Net Pay
    net_pay = gross_pay - total_deductions
    
    print(f"[PAYSLIP_CALC] Total deductions: {total_deductions:.2f}, Net pay: {net_pay:.2f}")
    
    # Build payslip data structure matching requirements
    payslip_data = {
        "employee": employee,
        "contract_wage": wage, # Full Basic
        "basic_pay": round(earned_basic_pay, 2), # Earned Basic
        "gross_pay": round(gross_pay, 2),
        "housing_allowance": housing_allowance,
        "transport_allowance": transport_allowance,
        "other_allowance": other_allowance,
        "total_allowances": round(total_allowances, 2),
        "taxable_gross_pay": round(gross_pay, 2),
        "net_pay": round(net_pay, 2),
        "allowances": [],
        "paid_days": round(paid_days, 2),
        "unpaid_days": round(unpaid_days, 2),
        "loss_of_pay": round(loss_of_pay_amount, 2), # Amount
        "lop_days": round(unpaid_days, 2), # Days
        "status": "draft",
        "overtime": round(overtime_hours, 2),
        "salary_advance_loan_recovery": salary_advance_loan_recovery,
        "deduction": round(total_deductions, 2), # Explicit deductions field
        "salary_advance": salary_advance,
        "bonus": bonus, 
        "total_deductions": round(total_deductions, 2), # Same as Deduction?
        "federal_tax": 0,
        "start_date": start_date,
        "end_date": end_date,
        "range": f"{start_date.strftime('%b %d %Y')} - {end_date.strftime('%b %d %Y')}",
        "present_days": present_days,
        "half_day_present": half_day_present,
        "absent_days": absent_days,
        "leave_full_days": leave_full_days,
        "leave_half_days": leave_half_days,
        "weekly_off_days": weekly_off_days,
        "holiday_days": holiday_days,
        "total_working_days": total_working_days,
        "employee_id": employee.badge_id,
        "employee_name": employee.get_full_name(),
        "department": department_name
    }
    
    # Convert to JSON format for PayHeadData
    # We need to ensure all keys required by template are present
    data_to_json = payslip_data.copy()
    data_to_json["employee"] = employee.id
    data_to_json["start_date"] = start_date.strftime("%Y-%m-%d")
    data_to_json["end_date"] = end_date.strftime("%Y-%m-%d")
    
    payslip_data["json_data"] = json.dumps(data_to_json, default=str)
    payslip_data["pay_head_data"] = data_to_json
    payslip_data["installments"] = []
    
    print(f"[PAYSLIP_CALC] Calculation completed successfully for Employee: {employee.badge_id}")
    
    return payslip_data


@login_required
@permission_required("payroll.add_payslip")
def payslip_import_info(request):
    """
    Renders the payslip import page with month/year selection and file upload.
    Similar to run_payroll.html UI.
    """
    # Get current year and month for default values
    today = date.today()
    try:
        current_year = int(request.GET.get('year', today.year))
    except (ValueError, TypeError):
        current_year = today.year
    
    current_month = today.month
    
    # Get all months with their status
    months_data = []
    for month_num in range(1, 13):
        month_name = calendar.month_name[month_num]
        # Check if payslips exist for this month
        month_start = date(current_year, month_num, 1)
        if month_num == 12:
            month_end = date(current_year, month_num, 31)
        else:
            month_end = date(current_year, month_num + 1, 1) - timedelta(days=1)
        
        payslips_count = Payslip.objects.filter(
            start_date__gte=month_start,
            end_date__lte=month_end
        ).count()
        
        status = "completed" if payslips_count > 0 else "pending"
        
        months_data.append({
            "month_num": month_num,
            "month_name": month_name,
            "year": current_year,
            "status": status,
            "date_range": f"{month_start.strftime('%b %d')} - {month_end.strftime('%b %d')}"
        })
    
    # Get employee count for selected company
    # Uses the same logic as payslip_import_template - get company from session or query param
    company_id = request.GET.get("company_id") or request.session.get("selected_company")
    
    # Count employees for the selected company
    # HorillaCompanyManager automatically scopes Employee.objects.all() by selected company
    employee_count = Employee.objects.all().count()
    
    context = {
        "current_year": current_year,
        "current_month": current_month,
        "months_data": months_data,
        "employee_count": employee_count,
    }
    
    return render(request, "payroll/payslip_import/payslip_import.html", context)


@login_required
@permission_required("payroll.add_payslip")
def payslip_import_template(request):
    """
    Download a payslip import template with component-level calculations.
    
    Correction:
    - Changed SUM to ROUND for Basic, Housing, Transport, and Other allowances.
    - Formula: =ROUND((Base_Amount / Days_In_Month) * Total_Paid_Days, 2)
    """
    company_id = request.GET.get("company_id") or request.session.get("selected_company")

    # --- 1. GET MONTH/YEAR & DAYS ---
    current_date = datetime.now()
    try:
        selected_month = int(request.GET.get("month", current_date.month))
        selected_year = int(request.GET.get("year", current_date.year))
        
        # Validate month is between 1 and 12
        if selected_month < 1 or selected_month > 12:
            selected_month = current_date.month
        
        # Validate year is reasonable (between 2000 and 2100)
        if selected_year < 2000 or selected_year > 2100:
            selected_year = current_date.year
            
    except (ValueError, TypeError):
        selected_month = current_date.month
        selected_year = current_date.year

    # Get exact days in month (28, 29, 30, or 31)
    # This value is used as the denominator in the formula
    # calendar.monthrange(year, month) returns (weekday_of_first_day, number_of_days)
    try:
        days_in_month = calendar.monthrange(selected_year, selected_month)[1]
    except (ValueError, TypeError):
        # Fallback to current month if calculation fails
        days_in_month = calendar.monthrange(current_date.year, current_date.month)[1]
        selected_month = current_date.month
        selected_year = current_date.year

    employees_qs = Employee.objects.all().select_related(
        "employee_work_info",
        "employee_work_info__department_id"
    )

    employees = employees_qs.values(
        "badge_id",
        "employee_first_name",
        "employee_last_name",
        "employee_work_info__basic_salary",
        "employee_work_info__housing_allowance",
        "employee_work_info__transport_allowance",
        "employee_work_info__other_allowance",
        "employee_work_info__department_id__department",
    )

    export_rows = []
    
    for emp in employees:
        emp_id = emp.get("badge_id") or ""
        first_name = emp.get("employee_first_name") or ""
        last_name = emp.get("employee_last_name") or ""
        emp_name = (f"{first_name} {last_name}").strip()
        dept_name = emp.get("employee_work_info__department_id__department") or ""
        
        # Base monthly amounts
        basic_pay = float(emp.get("employee_work_info__basic_salary") or 0)
        housing = float(emp.get("employee_work_info__housing_allowance") or 0)
        transport = float(emp.get("employee_work_info__transport_allowance") or 0)
        other = float(emp.get("employee_work_info__other_allowance") or 0)
        
        export_rows.append({
            # Static columns
            "Employee Id": emp_id,
            "Employee Name": emp_name,
            "Department": dept_name,
            # Base values (overwritten by formula in loop)
            "Basic Salary": basic_pay,
            "Housing Allowance": housing,
            "Transport Allowance": transport,
            "Other Allowance": other,
            
            # Default "Total Paid Days" to the full month days
            "Total Paid Days": days_in_month, 
            
            "Gross Salary": "", 
            "Loss of Pay": "",
            "salary_advance_loan_recovery": "",
            "Deduction": "",
            "Overtime": "",
            "salary_advance": "",
            "bonus": "",
            "Net Salary": "",
        })

    df = pd.DataFrame(export_rows)
    
    # Column Definitions
    static_cols = ["Employee Id", "Employee Name", "Department", "Basic Salary", "Housing Allowance", "Transport Allowance", "Other Allowance"]
    variable_cols = ["Total Paid Days", "Gross Salary", "Loss of Pay", "salary_advance_loan_recovery", "Deduction", "Overtime", "salary_advance", "bonus", "Net Salary"]

    file_name = f"payslip_import_template_{company_id or 'all'}_{selected_year}_{selected_month}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    writer = pd.ExcelWriter(response, engine="xlsxwriter")
    df.to_excel(writer, index=False, sheet_name="Payslips", startrow=2)
    
    workbook = writer.book
    worksheet = writer.sheets["Payslips"]
    
    # --- Formats ---
    fmt_green_header = workbook.add_format({'bold': True, 'bg_color': '#D8E4BC', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    fmt_green_data = workbook.add_format({'bg_color': '#D8E4BC', 'border': 1, 'align': 'left', 'valign': 'vcenter'})
    
    fmt_orange_header = workbook.add_format({'bold': True, 'bg_color': '#FCD5B4', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
    fmt_orange_data = workbook.add_format({'bg_color': '#FCD5B4', 'border': 1, 'align': 'left', 'valign': 'vcenter'})
    
    merged_header_fmt = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})

    # --- Header Setup ---
    worksheet.merge_range(0, 9, 0, 11, "DEDUCTION", merged_header_fmt)
    worksheet.merge_range(0, 12, 0, 14, "ADDITION", merged_header_fmt)
    
    for col_num, col_name in enumerate(df.columns.values):
        if col_name in static_cols:
            worksheet.write(2, col_num, col_name, fmt_green_header)
        else:
            worksheet.write(2, col_num, col_name, fmt_orange_header)
        worksheet.set_column(col_num, col_num, 18)
    
    # --- Data Writing & Formulas ---
    num_data_rows = len(df)
    for row_idx in range(num_data_rows):
        data_row = row_idx + 3
        excel_row = data_row + 1
        
        # Base values for formula construction
        base_basic = df.iloc[row_idx]['Basic Salary']
        base_housing = df.iloc[row_idx]['Housing Allowance']
        base_transport = df.iloc[row_idx]['Transport Allowance']
        base_other = df.iloc[row_idx]['Other Allowance']

        for col_num, col_name in enumerate(df.columns.values):
            cell_value = df.iloc[row_idx, col_num]
            cell_format = fmt_green_data if col_name in static_cols else fmt_orange_data
            
            # --- FIX: Using ROUND() instead of SUM() ---
            if col_name == "Basic Salary":
                formula = f"=ROUND(({base_basic}/{days_in_month})*H{excel_row}, 2)"
                worksheet.write_formula(data_row, col_num, formula, cell_format)
                
            elif col_name == "Housing Allowance":
                formula = f"=ROUND(({base_housing}/{days_in_month})*H{excel_row}, 2)"
                worksheet.write_formula(data_row, col_num, formula, cell_format)
                
            elif col_name == "Transport Allowance":
                formula = f"=ROUND(({base_transport}/{days_in_month})*H{excel_row}, 2)"
                worksheet.write_formula(data_row, col_num, formula, cell_format)
                
            elif col_name == "Other Allowance":
                formula = f"=ROUND(({base_other}/{days_in_month})*H{excel_row}, 2)"
                worksheet.write_formula(data_row, col_num, formula, cell_format)
                
            elif col_name == "Gross Salary":
                # Sum of the above calculated columns (D, E, F, G)
                gross_formula = f"=SUM(D{excel_row}:G{excel_row})"
                worksheet.write_formula(data_row, col_num, gross_formula, cell_format)
                
            elif col_name == "Net Salary":
                # Gross - Deductions + Additions
                net_formula = f"=I{excel_row}-SUM(J{excel_row}:L{excel_row})+SUM(M{excel_row}:O{excel_row})"
                worksheet.write_formula(data_row, col_num, net_formula, cell_format)
                
            else:
                # Standard write handles the integer 'days_in_month' correctly
                if pd.isna(cell_value) or cell_value == "" or (isinstance(cell_value, str) and cell_value.strip() == ""):
                    worksheet.write_blank(data_row, col_num, None, cell_format)
                elif isinstance(cell_value, (int, float)):
                    worksheet.write_number(data_row, col_num, float(cell_value), cell_format)
                else:
                    worksheet.write(data_row, col_num, str(cell_value), cell_format)

    # Final Formatting
    worksheet.set_row(0, 20)
    worksheet.set_row(1, 5) 
    worksheet.set_row(2, 25)

    writer.close()
    return response


@permission_required("payroll.change_payslip")
@permission_required('payroll.add_payslip', raise_exception=True)
def payslip_import_process(request):
    """
    Handles the POST request for uploading Excel file and automatically calculating payslips
    based ONLY on imported Excel data (no attendance calculations).
    
    Expected Excel columns:
    - Employee Id, Employee Name, Department, Basic Salary, Housing Allowance, Transport Allowance, Other Allowance,
    - Gross Salary, Total Paid Days, Loss of Pay, salary_advance_loan_recovery, Deduction,
    - Overtime, salary_advance, bonus, Net Salary
    
    Formula: Net Pay = (Gross Pay + Overtime + salary_advance + bonus) - (Loss of Pay + salary_advance_loan_recovery + Deduction)
    """
    
    if request.method != "POST":
        return JsonResponse({
            'status': 'error',
            'message': "Only POST method allowed."
        }, status=400)

    # Get month and year from request
    try:
        month = int(request.POST.get('month', date.today().month))
        year = int(request.POST.get('year', date.today().year))
    except (ValueError, TypeError):
        return JsonResponse({
            'status': 'error',
            'message': "Invalid month or year provided."
        }, status=400)

    # Calculate start and end dates for the selected month
    if month == 12:
        month_end = date(year, month, 31)
    else:
        month_end = date(year, month + 1, 1) - timedelta(days=1)
    
    start_date = date(year, month, 1)
    end_date = month_end
    
    # Check if file is uploaded
    if 'file' not in request.FILES:
        return JsonResponse({
            'status': 'error',
            'message': "No file uploaded. Please upload an Excel file."
        }, status=400)

    file = request.FILES['file']
    file_extension = file.name.split('.')[-1].lower()

    if file_extension not in ['xls', 'xlsx', 'csv']:
        return JsonResponse({
            'status': 'error',
            'message': "Unsupported file format. Please use .xls, .xlsx, or .csv"
        }, status=400)

    try:
        from io import BytesIO
        file_data = BytesIO(file.read())

        # Read Excel file - handle template format with merged headers
        header_row = 0
        if file_extension in ['xls', 'xlsx']:
            try:
                file_data.seek(0)
                preview = pd.read_excel(file_data, nrows=4, header=None)
                file_data.seek(0)
                
                # Check if row 0 contains merged header keywords
                row0_values = [str(v).upper().strip() for v in preview.iloc[0].values if pd.notna(v)]
                row0_text = ' '.join(row0_values)
                has_merged_headers = any(keyword in row0_text for keyword in ['DEDUCTION', 'ADDITION'])
                
                # Check if row 2 contains column header keywords
                row2_values = [str(v).upper().strip() for v in preview.iloc[2].values if pd.notna(v)]
                row2_text = ' '.join(row2_values)
                has_column_headers = any(keyword in row2_text for keyword in ['EMPLOYEE ID', 'EMPLOYEE_NAME', 'EMPLOYEEID', 'BASIC SALARY', 'BASIC_SALARY', 'GROSS SALARY'])
                
                if has_merged_headers and has_column_headers and len(preview) > 2:
                    header_row = 2
                    print(f"[IMPORT] Detected template format, using row {header_row} as header")
            except Exception as e:
                print(f"[IMPORT] Header detection failed: {e}, using default header row 0")
            
            df = pd.read_excel(file_data, header=header_row)
        else:
            df = pd.read_csv(file_data)

        # Normalize column names for matching
        df.columns = (
            df.columns.astype(str)
            .str.strip()
            .str.lower()
            .str.replace(" ", "_")
            .str.replace("-", "_")
        )

        # Check for employee_id column
        employee_id_col = None
        for col in df.columns:
            normalized_col = col.strip().lower().replace(" ", "_").replace("-", "_")
            if normalized_col in ['employee_id', 'employeeid', 'emp_id', 'badge_id', 'badgeid']:
                employee_id_col = col
                break
        
        if not employee_id_col:
            if 'employee_id' not in df.columns:
                return JsonResponse({
                    'status': 'error',
                    'message': f"Missing required column: 'Employee Id'. Found columns: {list(df.columns)}"
                }, status=400)
            employee_id_col = 'employee_id'
        
        if employee_id_col != 'employee_id':
            df = df.rename(columns={employee_id_col: 'employee_id'})

        # Check for basic_pay/basic_salary column
        wage_column = None
        for col in df.columns:
            normalized_col = col.strip().lower().replace(" ", "_").replace("-", "_")
            if normalized_col in ['basic_pay', 'basic_salary', 'wage']:
                wage_column = col
                break
        
        if not wage_column:
            return JsonResponse({
                'status': 'error',
                'message': f"Missing required column: 'Basic Salary' or 'Basic Pay'. Found columns: {list(df.columns)}"
            }, status=400)
        
        if wage_column != 'basic_pay':
            df = df.rename(columns={wage_column: 'basic_pay'})

        # Remove label rows
        label_texts = [
            'static components for every payroll month',
            'variable components to be changed/entered every month',
            'static components',
            'variable components'
        ]
        
        rows_to_drop = []
        for idx in df.index:
            row_values = ' '.join([str(val).lower().strip() for val in df.loc[idx].values if pd.notna(val)])
            if any(label_text in row_values for label_text in label_texts):
                rows_to_drop.append(idx)
        
        if rows_to_drop:
            df = df.drop(rows_to_drop).reset_index(drop=True)

        errors = []
        total_imported = 0

        with transaction.atomic():
            for index, row in df.iterrows():
                row_no = header_row + index + 2  # Excel row number (1-indexed)

                try:
                    employee_id = str(row['employee_id']).strip()
                    if not employee_id or employee_id.lower() in ['nan', 'none', '']:
                        continue
                    
                    # Skip label rows
                    if any(label_text in employee_id.lower() for label_text in label_texts):
                        continue

                    # Get employee
                    try:
                        employee = Employee.objects.get(badge_id=employee_id)
                    except Employee.DoesNotExist:
                        raise ValueError(f"Employee with badge ID '{employee_id}' not found.")

                    # Get ALL values from Excel ONLY (no fallbacks to system values)
                    # Column was renamed to 'basic_pay' earlier, so access it directly
                    monthly_basic_pay = float(np.nan_to_num(row.get('basic_pay', 0)))
                    if monthly_basic_pay == 0:
                        # Fallback: try 'basic_salary' in case rename didn't work
                        monthly_basic_pay = float(np.nan_to_num(row.get('basic_salary', 0)))
                    # Read Department from Excel
                    department_excel = str(row.get('department', '')).strip() if pd.notna(row.get('department', '')) else ''
                    housing_allowance = float(np.nan_to_num(row.get('housing_allowance', 0)))
                    transport_allowance = float(np.nan_to_num(row.get('transport_allowance', 0)))
                    other_allowance = float(np.nan_to_num(row.get('other_allowance', 0)))
                    gross_pay_excel = float(np.nan_to_num(row.get('gross_pay', row.get('gross_salary', 0))))
                    paid_days = float(np.nan_to_num(row.get('total_paid_days', 0)))
                    loss_of_pay = float(np.nan_to_num(row.get('lop', row.get('loss_of_pay', 0))))
                    salary_advance_loan_recovery = float(np.nan_to_num(row.get('salary_advance_loan_recovery', 0)))
                    deduction = float(np.nan_to_num(row.get('deduction', 0)))
                    overtime = float(np.nan_to_num(row.get('overtime', 0)))
                    salary_advance = float(np.nan_to_num(row.get('salary_advance', 0)))
                    bonus = float(np.nan_to_num(row.get('bonus', 0)))
                    
                    # Calculate Basic Pay: (Monthly Basic Pay / Calendar month days) × Paid Days from Excel
                    # BYPASS CONTRACT: Use Excel values directly, no contract dependency
                    calc_start_date = start_date
                    calc_end_date = end_date
                    
                    # Calculate calendar days in the period (for paid_days validation)
                    days_in_period = (calc_end_date - calc_start_date).days + 1
                    
                    # CRITICAL: Use calendar month days for per-day calculation (30/31/28/29)
                    # Always use the actual calendar month days, not the period days
                    from calendar import monthrange
                    # Use the month of start_date for calendar days calculation
                    calendar_month_days = monthrange(calc_start_date.year, calc_start_date.month)[1]
                    
                    # Calculate per-day basic pay using calendar month days
                    per_day_basic_pay = monthly_basic_pay / calendar_month_days if calendar_month_days > 0 else 0
                    
                    # CRITICAL: Calculate paid_days and unpaid_days using the same logic as monthly_computation
                    # This ensures consistency between generated payslips and imported payslips
                    # Get month data and leave data for proper calculation (same as monthly_computation)
                    month_data = months_between_range(monthly_basic_pay, calc_start_date, calc_end_date)
                    leave_data = get_leaves(employee, calc_start_date, calc_end_date)
                    
                    # Calculate half-day leaves (same logic as monthly_computation)
                    date_range_list = get_date_range(calc_start_date, calc_end_date)
                    if apps.is_installed("leave"):
                        start_date_leaves = (
                            employee.leaverequest_set.filter(
                                leave_type_id__payment="unpaid",
                                start_date__in=date_range_list,
                                status="approved",
                            )
                            .exclude(start_date_breakdown="full_day")
                            .count()
                        )
                        end_date_leaves = (
                            employee.leaverequest_set.filter(
                                leave_type_id__payment="unpaid",
                                end_date__in=date_range_list,
                                status="approved",
                            )
                            .exclude(end_date_breakdown="full_day")
                            .exclude(start_date=F("end_date"))
                            .count()
                        )
                    else:
                        start_date_leaves = 0
                        end_date_leaves = 0
                    
                    half_day_leaves_between_period_on_start_date = start_date_leaves
                    half_day_leaves_between_period_on_end_date = end_date_leaves
                    
                    unpaid_half_leaves = (
                        half_day_leaves_between_period_on_start_date
                        + half_day_leaves_between_period_on_end_date
                    ) * 0.5
                    
                    # Calculate unpaid leaves (excluding half-day adjustments) - same as monthly_computation
                    calculated_unpaid_leaves = abs(leave_data["unpaid_leaves"] - unpaid_half_leaves)
                    
                    # Get working days from month_data (same as monthly_computation)
                    if month_data and isinstance(month_data, list) and len(month_data) > 0:
                        working_days_on_period = month_data[0]["working_days_on_period"]
                    else:
                        # Fallback: use calendar days if month_data is not available
                        working_days_on_period = days_in_period
                    
                    # Calculate paid_days using the same logic as monthly_computation
                    # paid_days = working_days_on_period - unpaid_leaves
                    calculated_paid_days = working_days_on_period - calculated_unpaid_leaves
                    
                    # Ensure paid_days is never negative
                    if calculated_paid_days < 0:
                        calculated_paid_days = 0
                    
                    # Use Excel paid_days if provided and valid, otherwise use calculated value
                    # This allows Excel to override, but ensures we have proper unpaid_days calculation
                    excel_paid_days = paid_days  # Store original Excel value
                    if excel_paid_days > 0:
                        # Excel provided paid_days - use it, but calculate unpaid_days from working days
                        unpaid_days = working_days_on_period - excel_paid_days if working_days_on_period > 0 else calculated_unpaid_leaves
                        if unpaid_days < 0:
                            unpaid_days = 0
                        # Use Excel paid_days as provided
                        paid_days = excel_paid_days
                    else:
                        # Excel didn't provide paid_days - use calculated values (same as generated payslips)
                        unpaid_days = calculated_unpaid_leaves
                        paid_days = calculated_paid_days
                    
                    # Final Basic Pay = (Monthly Basic Pay / Calendar month days) × Paid Days
                    # Uses calculated paid_days (from leaves) or Excel paid_days if provided
                    basic_pay = monthly_basic_pay
                    
                    # Use Excel gross_pay if provided, otherwise calculate from basic + allowances
                    if gross_pay_excel > 0:
                        gross_pay = gross_pay_excel
                        print("line3940", gross_pay)
                    else:
                        gross_pay = basic_pay + housing_allowance + transport_allowance + other_allowance
                        
                        print("line3943", gross_pay)
                        print("line3944", basic_pay)
                        print("line3945", housing_allowance)
                        print("line3946", transport_allowance)
                        print("line3947", other_allowance)
                        print("line3948", gross_pay_excel)
                        print("line3949", paid_days)
                        print("line3950", loss_of_pay)
                        print("line3951", salary_advance_loan_recovery)
                        print("line3952", deduction)
                        print("line3953", overtime)
                        print("line3954", salary_advance)
                        print("line3955", bonus)
                    
                    # Apply Excel formula: Net Pay = (Gross Pay + Overtime + salary_advance + bonus) - (Loss of Pay + salary_advance_loan_recovery + Deduction)
                    net_pay = round(
                        (gross_pay + overtime + salary_advance + bonus) - 
                        (loss_of_pay + salary_advance_loan_recovery + deduction),
                        2
                    )
                    
                    # Calculate total allowances
                    total_allowances = housing_allowance + transport_allowance + other_allowance
                    
                    # Calculate total deductions
                    total_deductions = loss_of_pay + salary_advance_loan_recovery + deduction
                    
                    # Use Excel monthly basic pay as contract_wage (bypass contract)
                    contract_wage = monthly_basic_pay

                    # Check if payslip already exists
                    existing_payslip = Payslip.objects.filter(
                        employee_id=employee,
                        start_date=calc_start_date,
                        end_date=calc_end_date
                    ).first()
                    
                    # Fetch and process Deduction components for the employee
                    # Get applicable deductions for the employee in the date range
                    applicable_deductions = (
                        Deduction.objects.filter(
                            specific_employees=employee,
                        )
                        | Deduction.objects.filter(
                            is_condition_based=True,
                        ).exclude(exclude_employees=employee)
                        | Deduction.objects.filter(
                            include_active_employees=True,
                        ).exclude(exclude_employees=employee)
                    )
                    
                    # Filter deductions that are applicable based on conditions
                    employee_deductions = []
                    for deduction_component in applicable_deductions:
                        applicable = True
                        if deduction_component.is_condition_based:
                            conditions = list(
                                deduction_component.other_conditions.values_list(
                                    "field", "condition", "value"
                                )
                            )
                            conditions.append(
                                (
                                    deduction_component.field,
                                    deduction_component.condition,
                                    deduction_component.value.lower().replace(" ", "_"),
                                )
                            )
                            for field, operator, value in conditions:
                                val = dynamic_attr(employee, field)
                                if val is None or not operator_mapping.get(operator)(
                                    val, type(val)(value)
                                ):
                                    applicable = False
                                    break
                        if applicable:
                            employee_deductions.append(deduction_component)
                    
                    # Process deductions and categorize them
                    basic_pay_deductions_list = []
                    gross_pay_deductions_list = []
                    pretax_deductions_list = []
                    post_tax_deductions_list = []
                    tax_deductions_list = []
                    net_deductions_list = []
                    
                    # Process deductions using update_compensation_deduction to get actual amounts
                    # Update basic_pay with basic_pay deductions
                    updated_basic_pay_data = update_compensation_deduction(
                        employee, basic_pay, "basic_pay", calc_start_date, calc_end_date
                    )
                    basic_pay = updated_basic_pay_data["compensation_amount"]
                    basic_pay_deductions_list = updated_basic_pay_data["deductions"]
                    
                    # Update gross_pay with gross_pay deductions
                    updated_gross_pay_data = update_compensation_deduction(
                        employee, gross_pay, "gross_pay", calc_start_date, calc_end_date
                    )
                    gross_pay = updated_gross_pay_data["compensation_amount"]
                    gross_pay_deductions_list = updated_gross_pay_data["deductions"]
                    
                    # Recalculate net_pay after processing deductions
                    # Net Pay = (Gross Pay + Overtime + salary_advance + bonus) - (Loss of Pay + salary_advance_loan_recovery + Deduction)
                    net_pay = round(
                        (gross_pay + overtime + salary_advance + bonus) - 
                        (loss_of_pay + salary_advance_loan_recovery + deduction),
                        2
                    )
                    
                    # Update net_pay with net_pay deductions
                    updated_net_pay_data = update_compensation_deduction(
                        employee, net_pay, "net_pay", calc_start_date, calc_end_date
                    )
                    net_pay = updated_net_pay_data["compensation_amount"]
                    net_deductions_list = updated_net_pay_data["deductions"]
                    
                    # CRITICAL: Clean deduction lists to ensure only JSON-serializable data
                    # Remove any model instances and keep only dicts with title, amount, id
                    def clean_deduction_list(deduction_list):
                        """Remove non-JSON-serializable objects from deduction list"""
                        if not isinstance(deduction_list, list):
                            return []
                        cleaned_list = []
                        for d in deduction_list:
                            # Skip if it's a model instance (has Django model attributes)
                            if hasattr(d, '_meta') or hasattr(d, 'save'):
                                # It's a model instance, skip it or convert to dict
                                continue
                            if isinstance(d, dict):
                                # Only keep JSON-serializable fields
                                cleaned_d = {
                                    "title": str(d.get("title", "")),
                                    "amount": float(d.get("amount", 0))
                                }
                                # Only add id if it exists and is serializable
                                if "id" in d:
                                    try:
                                        cleaned_d["id"] = int(d["id"])
                                    except (ValueError, TypeError):
                                        pass
                                # Remove any non-serializable fields (like "component")
                                for key in list(cleaned_d.keys()):
                                    if key not in ["title", "amount", "id"]:
                                        del cleaned_d[key]
                                cleaned_list.append(cleaned_d)
                        return cleaned_list
                    
                    # Clean all deduction lists
                    basic_pay_deductions_list = clean_deduction_list(basic_pay_deductions_list)
                    gross_pay_deductions_list = clean_deduction_list(gross_pay_deductions_list)
                    pretax_deductions_list = clean_deduction_list(pretax_deductions_list)
                    post_tax_deductions_list = clean_deduction_list(post_tax_deductions_list)
                    tax_deductions_list = clean_deduction_list(tax_deductions_list)
                    net_deductions_list = clean_deduction_list(net_deductions_list)
                    
                    # Recalculate total_deductions to include component deductions
                    component_deduction_total = sum(
                        d.get("amount", 0) for d in (
                            basic_pay_deductions_list + 
                            gross_pay_deductions_list + 
                            pretax_deductions_list + 
                            post_tax_deductions_list + 
                            tax_deductions_list + 
                            net_deductions_list
                        )
                    )
                    total_deductions = loss_of_pay + salary_advance_loan_recovery + deduction + component_deduction_total
                    
                    # Build pay_head_data with ALL Excel values and component deductions
                    pay_head_data = {
                        'paid_days': round(paid_days, 2),
                        'unpaid_days': round(unpaid_days, 2) if unpaid_days > 0 else 0,
                        'lop_days': round(unpaid_days, 2) if unpaid_days > 0 else 0,
                        'loss_of_pay': round(loss_of_pay, 2),
                        'basic_pay': round(basic_pay, 2),
                        'housing_allowance': round(housing_allowance, 2),
                        'transport_allowance': round(transport_allowance, 2),
                        'other_allowance': round(other_allowance, 2),
                        'total_allowances': round(total_allowances, 2),
                        'salary_advance_loan_recovery': round(salary_advance_loan_recovery, 2),
                        'deduction': round(deduction, 2),
                        'salary_advance': round(salary_advance, 2),
                        'bonus': round(bonus, 2),
                        'overtime': round(overtime, 2),
                        'total_deductions': round(total_deductions, 2),
                        'gross_pay': round(gross_pay, 2),
                        'net_pay': round(net_pay, 2),
                        'contract_wage': round(contract_wage, 2),
                        'monthly_basic_pay': round(monthly_basic_pay, 2),  # Store monthly basic for reference
                        'department': department_excel,  # Store Department from Excel
                        'is_imported': True,  # Mark as imported
                        # Include component deductions
                        'allowances': [],
                        'pretax_deductions': pretax_deductions_list,
                        'posttax_deductions': post_tax_deductions_list,
                        'post_tax_deductions': post_tax_deductions_list,
                        'basic_pay_deductions': basic_pay_deductions_list,
                        'gross_pay_deductions': gross_pay_deductions_list,
                        'tax_deductions': tax_deductions_list,
                        'net_deductions': net_deductions_list,
                    }
                    
                    print(f"[IMPORT] Row {row_no} - Employee: {employee_id}")
                    print(f"[IMPORT]   Monthly Basic Pay: {monthly_basic_pay}, Calendar Month Days: {calendar_month_days}, Period Days: {days_in_period}, Paid Days: {paid_days}")
                    print(f"[IMPORT]   Per-day Basic Pay: {per_day_basic_pay:.10f}, Final Basic Pay: {basic_pay:.2f}")
                    print(f"[IMPORT]   Overtime: {overtime}, Salary Advance Loan Recovery: {salary_advance_loan_recovery}, Deduction: {deduction}")
                    print(f"[IMPORT]   Gross Pay: {gross_pay:.2f}, Net Pay: {net_pay:.2f}")
                    # Log component deductions
                    if employee_deductions:
                        print(f"[IMPORT]   Component Deductions: {len(employee_deductions)} found")
                        for comp in employee_deductions:
                            print(f"[IMPORT]     - Component: {comp}")

                    # Create or update payslip
                    if existing_payslip:
                        existing_payslip.contract_wage = round(contract_wage, 2)
                        existing_payslip.basic_pay = round(basic_pay, 2)
                        existing_payslip.gross_pay = round(gross_pay, 2)
                        existing_payslip.net_pay = round(net_pay, 2)
                        existing_payslip.deduction = round(total_deductions, 2)
                        existing_payslip.pay_head_data = pay_head_data
                        if existing_payslip.status != 'paid':
                            existing_payslip.status = 'draft'
                        existing_payslip.save()
                        print(f"[IMPORT] Updated payslip for {employee} - Net Pay: {net_pay:.2f}")
                    else:
                        Payslip.objects.create(
                            employee_id=employee,
                            start_date=calc_start_date,
                            end_date=calc_end_date,
                            contract_wage=round(contract_wage, 2),
                            basic_pay=round(basic_pay, 2),
                            gross_pay=round(gross_pay, 2),
                            net_pay=round(net_pay, 2),
                            deduction=round(total_deductions, 2),
                            pay_head_data=pay_head_data,
                            status='draft'
                        )
                        print(f"[IMPORT] Created payslip for {employee} - Net Pay: {net_pay:.2f}")

                    total_imported += 1

                except Exception as e:
                    import traceback
                    traceback.print_exc()
                    errors.append(f"Row {row_no}: {str(e)}")

            if errors:
                error_detail = "\n".join(errors[:10]) + ("..." if len(errors) > 10 else "")
                return JsonResponse({
                    'status': 'error',
                    'message': f"Import completed with {len(errors)} errors:\n{error_detail}",
                    'imported': total_imported,
                    'errors': len(errors)
                }, status=200)

        return JsonResponse({
            'status': 'success',
            'message': f"Successfully generated {total_imported} payslips for {calendar.month_name[month]} {year}.",
            'imported': total_imported
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': f"Import failed: {str(e)}"
        }, status=400)


def calculate_payslip_values(payslip):
    """
    Calculate all payslip values including database allowances and deductions.
    This function ensures consistency between table view and export functionality.
    
    Returns a dict with all calculated values:
    - basic_pay, housing_allowance, transport_allowance, other_allowance
    - gross_pay (including all database allowances)
    - paid_days, unpaid_days, loss_of_pay
    - overtime, salary_advance, bonus
    - salary_advance_loan_recovery, deduction (including component deductions)
    - net_pay
    
    Also updates payslip.pay_head_data with calculated values.
    """
    employee = payslip.employee_id
    if not employee:
        return None
    
    # CRITICAL: Extract and process pay_head_data
    pay_head = payslip.pay_head_data
    if pay_head is None:
        pay_head = {}
    elif isinstance(pay_head, str):
        try:
            pay_head = json.loads(pay_head)
        except (json.JSONDecodeError, TypeError):
            pay_head = {}
    if not isinstance(pay_head, dict):
        pay_head = {}
    
    # Create a copy to avoid modifying the original
    pay_head = dict(pay_head)
    
    # Calculate paid_days and unpaid_days
    unpaid_days_value = pay_head.get("unpaid_days")
    if unpaid_days_value is None:
        unpaid_days_value = pay_head.get("lop_days")
    if unpaid_days_value is None:
        unpaid_days_value = 0
    
    try:
        unpaid_days = round(float(unpaid_days_value), 2)
    except (ValueError, TypeError):
        unpaid_days = 0
    
    # Calculate total days in period
    days_in_period = 0
    if payslip.start_date and payslip.end_date:
        days_in_period = (payslip.end_date - payslip.start_date).days + 1
    
    # Extract paid_days and validate/correct if needed
    paid_days_value = pay_head.get("paid_days")
    if paid_days_value is None or paid_days_value == "":
        if days_in_period > 0 and unpaid_days > 0:
            paid_days_value = days_in_period - unpaid_days
        else:
            paid_days_value = days_in_period if days_in_period > 0 else 0
    else:
        try:
            paid_days_value = float(paid_days_value)
            if paid_days_value == 0 and unpaid_days > 0 and days_in_period > 0:
                paid_days_value = days_in_period - unpaid_days
            elif days_in_period > 0 and unpaid_days > 0:
                total_calculated = paid_days_value + unpaid_days
                if total_calculated > days_in_period:
                    paid_days_value = days_in_period - unpaid_days
        except (ValueError, TypeError):
            if days_in_period > 0 and unpaid_days > 0:
                paid_days_value = days_in_period - unpaid_days
            else:
                paid_days_value = 0
    
    try:
        paid_days = round(float(paid_days_value), 2)
    except (ValueError, TypeError):
        paid_days = 0
    
    # Extract overtime, salary_advance, and bonus
    overtime = float(pay_head.get("overtime", 0) or 0)
    salary_advance = float(pay_head.get("salary_advance", 0) or 0)
    bonus = float(pay_head.get("bonus", 0) or 0)
    
    # Also check allowances list for bonus if not in pay_head_data directly
    if bonus == 0 and isinstance(pay_head.get("allowances"), list):
        for allowance in pay_head.get("allowances", []):
            if isinstance(allowance, dict) and "bonus" in str(allowance.get("title", "")).lower():
                bonus += float(allowance.get("amount", 0) or 0)
    
    # Extract all values with proper fallbacks
    basic_pay = round(float(pay_head.get("basic_pay", payslip.basic_pay or 0)), 2)
    housing_allowance = round(float(pay_head.get("housing_allowance", payslip.housing_allowance or 0)), 2)
    transport_allowance = round(float(pay_head.get("transport_allowance", payslip.transport_allowance or 0)), 2)
    other_allowance = round(float(pay_head.get("other_allowance", payslip.other_allowance or 0)), 2)
    
    # CRITICAL: Check if payslip is imported (needed before calculations)
    is_imported = pay_head.get("is_imported", False)
    
    # CRITICAL: Fetch allowances from database (same as view_created_payslip)
    allowances_queryset = Allowance.objects.filter(
        specific_employees=employee,
        only_show_under_employee=True
    )
    
    # Filter by date if one_time_date is set and falls within payslip period
    if payslip.start_date and payslip.end_date:
        allowances_queryset = allowances_queryset.filter(
            Q(one_time_date__isnull=True) | 
            Q(one_time_date__gte=payslip.start_date, one_time_date__lte=payslip.end_date)
        )
    
    # Process allowances from database
    db_allowances_total = 0
    other_allowances_total = 0
    
    for allowance_component in allowances_queryset:
        # Calculate amount
        amount = 0
        if allowance_component.is_fixed:
            amount = float(allowance_component.amount or 0)
        else:
            # For non-fixed, calculate based on rate and basic pay
            rate = float(allowance_component.rate or 0)
            amount = (basic_pay * rate) / 100
        
        if amount > 0:
            title_lower = allowance_component.title.lower()
            # Check if it's overtime, salary_advance, or bonus (these are handled separately)
            if "overtime" in title_lower:
                overtime += amount
            elif "salary advance" in title_lower or "advanced salary" in title_lower:
                salary_advance += amount
            elif "bonus" in title_lower:
                bonus += amount
            else:
                # Other allowances - include in gross_pay
                other_allowances_total += amount
                db_allowances_total += amount
    
    # CRITICAL: Use unified gross_pay calculation
    gross_pay = round(
        basic_pay + 
        housing_allowance + 
        transport_allowance + 
        other_allowance + 
        db_allowances_total, 2
    )
    
    # CRITICAL: loss_of_pay should only be shown for imported payslips OR inline-edited payslips
    has_inline_edit_loss_of_pay = (
        pay_head.get("loss_of_pay") is not None or 
        pay_head.get("lop") is not None
    )
    
    # Only extract/calculate loss_of_pay if payslip is imported OR inline-edited
    loss_of_pay_amount = 0
    if is_imported or has_inline_edit_loss_of_pay:
        loss_of_pay_amount = pay_head.get("loss_of_pay")
        if loss_of_pay_amount is None:
            loss_of_pay_amount = pay_head.get("lop", 0)
        if loss_of_pay_amount is None:
            loss_of_pay_amount = 0
        
        try:
            loss_of_pay_amount = round(float(loss_of_pay_amount), 2)
        except (ValueError, TypeError):
            loss_of_pay_amount = 0
    
    # Variable components with fallbacks
    salary_advance_loan_recovery = round(float(pay_head.get("salary_advance_loan_recovery", 0) or 0), 2)
    deduction = round(float(pay_head.get("deduction", payslip.deduction or 0) or 0), 2)
    
    # CRITICAL: Fetch component deductions from database
    deductions_queryset = Deduction.objects.filter(
        specific_employees=employee,
        only_show_under_employee=True
    )
    
    # Filter by date if one_time_date is set and falls within payslip period
    if payslip.start_date and payslip.end_date:
        deductions_queryset = deductions_queryset.filter(
            Q(one_time_date__isnull=True) | 
            Q(one_time_date__gte=payslip.start_date, one_time_date__lte=payslip.end_date)
        )
    
    # Process component deductions
    component_deduction_total = 0
    
    # Process gross_pay deductions (these affect gross_pay calculation)
    gross_pay_deduction_total = 0
    for deduction_component in deductions_queryset:
        if deduction_component.update_compensation == "gross_pay":
            amount = 0
            if deduction_component.is_fixed:
                amount = float(deduction_component.amount or 0)
            else:
                rate = float(deduction_component.rate or 0)
                amount = (gross_pay * rate) / 100
            if amount > 0:
                gross_pay_deduction_total += amount
    
    # Apply gross_pay deductions to gross_pay
    gross_pay = round(gross_pay - gross_pay_deduction_total, 2)
    
    # Process other deductions (pretax, post_tax, tax, net_pay)
    for deduction_component in deductions_queryset:
        if deduction_component.update_compensation:
            continue
        
        amount = 0
        if deduction_component.is_fixed:
            amount = float(deduction_component.amount or 0)
        else:
            if deduction_component.based_on == "basic_pay":
                base_amount = basic_pay
            elif deduction_component.based_on == "gross_pay":
                base_amount = gross_pay
            else:
                base_amount = basic_pay
            
            rate = float(deduction_component.rate or 0)
            amount = (base_amount * rate) / 100
        
        if amount > 0:
            component_deduction_total += amount
    
    # Process net_pay deductions (calculated after gross_pay deductions)
    initial_net_pay = round((gross_pay + overtime + salary_advance + bonus) - (loss_of_pay_amount + salary_advance_loan_recovery + deduction + component_deduction_total), 2)
    
    net_pay_deduction_total = 0
    for deduction_component in deductions_queryset:
        if deduction_component.update_compensation == "net_pay":
            amount = 0
            if deduction_component.is_fixed:
                amount = float(deduction_component.amount or 0)
            else:
                rate = float(deduction_component.rate or 0)
                amount = (initial_net_pay * rate) / 100
            if amount > 0:
                net_pay_deduction_total += amount
                component_deduction_total += amount
    
    # CRITICAL: Include component deductions in total deduction calculation
    total_deductions_including_components = round(
        loss_of_pay_amount + salary_advance_loan_recovery + deduction + component_deduction_total, 
        2
    )
    
    # CRITICAL: Unified Net Pay calculation
    net_pay = round(initial_net_pay - net_pay_deduction_total, 2)
    
    # Update pay_head_data with calculated values
    pay_head["paid_days"] = paid_days
    pay_head["unpaid_days"] = unpaid_days
    pay_head["lop_days"] = unpaid_days
    pay_head["loss_of_pay"] = loss_of_pay_amount
    pay_head["lop"] = loss_of_pay_amount
    pay_head["overtime"] = overtime
    pay_head["salary_advance"] = salary_advance
    pay_head["bonus"] = bonus
    pay_head["gross_pay"] = gross_pay
    pay_head["net_pay"] = net_pay
    pay_head["deduction"] = total_deductions_including_components
    pay_head["salary_advance_loan_recovery"] = salary_advance_loan_recovery
    
    # Update payslip object's pay_head_data
    payslip.pay_head_data = pay_head
    payslip.calculated_paid_days = paid_days
    payslip.calculated_unpaid_days = unpaid_days
    payslip.calculated_loss_of_pay = loss_of_pay_amount
    
    return {
        "basic_pay": basic_pay,
        "housing_allowance": housing_allowance,
        "transport_allowance": transport_allowance,
        "other_allowance": other_allowance,
        "gross_pay": gross_pay,
        "paid_days": paid_days,
        "unpaid_days": unpaid_days,
        "loss_of_pay": loss_of_pay_amount,
        "overtime": overtime,
        "salary_advance": salary_advance,
        "bonus": bonus,
        "salary_advance_loan_recovery": salary_advance_loan_recovery,
        "deduction": total_deductions_including_components,
        "net_pay": net_pay,
    }


@login_required
@permission_required("payroll.change_payslip")
def payslip_import_download(request):
    """
    Download payslips as Excel file using filters from URL parameters.
    Company-wise data is separated into different sheets.
    Data matches what's shown in the filtered payslip table view.
    """
    # CRITICAL: Use PayslipFilter to respect all filters from request (including company, date, employee filters)
    # This ensures downloaded data matches the on-screen filtered payroll data
    payslips = PayslipFilter(request.GET).qs.select_related(
        'employee_id', 
        'employee_id__employee_work_info', 
        'employee_id__employee_work_info__department_id',
        'employee_id__employee_work_info__company_id'
    )

    # CRITICAL: Evaluate queryset to list to ensure all payslips are fetched
    # This prevents lazy evaluation issues that might cause only partial data to be processed
    payslips_list = list(payslips)
    print(f"[PAYSLIP_IMPORT_DOWNLOAD] Found {len(payslips_list)} payslips to export")

    # Group payslips by company
    company_payslips = defaultdict(list)
    
    for payslip in payslips_list:
        try:
            employee = payslip.employee_id
            if not employee:
                print(f"[PAYSLIP_IMPORT_DOWNLOAD] Skipping payslip {payslip.id} - no employee")
                continue
                
            company = None
            if hasattr(employee, 'employee_work_info') and employee.employee_work_info:
                company = getattr(employee.employee_work_info, 'company_id', None)
            
            # Use company name as key, or "Unknown Company" if no company
            company_name = company.company if company else "Unknown Company"
            company_payslips[company_name].append(payslip)
        except Exception as e:
            print(f"[PAYSLIP_IMPORT_DOWNLOAD] Error processing payslip {payslip.id}: {e}")
            continue

    # Generate filename based on current date or filter parameters
    today = date.today()
    file_name = f"Payslip_excel_{today.strftime('%Y-%m-%d')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'

    writer = pd.ExcelWriter(response, engine="xlsxwriter")
    workbook = writer.book
    
    # Static components header format (green)
    static_header_fmt = workbook.add_format({
        'bold': True, 
        'bg_color': '#D8E4BC',  # Light green
        'border': 1,
        'align': 'center'
    })
    
    # Variable components header format (orange)
    variable_header_fmt = workbook.add_format({
        'bold': True, 
        'bg_color': '#FCD5B4',  # Orange
        'border': 1,
        'align': 'center'
    })
    
    # Static components data format (green background)
    static_data_fmt = workbook.add_format({
        'bg_color': '#D8E4BC',  # Light green
        'border': 1
    })
    
    # Variable components data format (orange background)
    variable_data_fmt = workbook.add_format({
        'bg_color': '#FCD5B4',  # Orange
        'border': 1
    })
    
    # Static columns: Employee Id, Employee Name, Department, Basic Pay, Housing Allowance, Transport Allowance, Other Allowance, Gross Pay
    static_cols = ["Employee Id", "Employee Name", "Department", "Basic Pay", "Housing Allowance", "Transport Allowance", "Other Allowance", "Gross Pay"]
    # Variable columns: Total Paid Days, LOP, Overtime, salary_advance_loan_recovery, Deduction, salary_advance, bonus, Net Pay
    variable_cols = ["Total Paid Days", "LOP", "Overtime", "salary_advance_loan_recovery", "Deduction", "salary_advance", "bonus", "Net Pay"]
    
    label_fmt_green = workbook.add_format({'bg_color': '#90EE90', 'bold': True})
    label_fmt_orange = workbook.add_format({'bg_color': '#FFA500', 'bold': True})
    
    # Track if any sheets were created
    sheets_created = False
    
    # Process each company separately
    for company_name, company_payslip_list in company_payslips.items():
        export_rows = []
        print(f"[PAYSLIP_IMPORT_DOWNLOAD] Processing company: {company_name} with {len(company_payslip_list)} payslips")
        
        for payslip in company_payslip_list:
            try:
                employee = payslip.employee_id
                if not employee:
                    print(f"[PAYSLIP_IMPORT_DOWNLOAD] Skipping payslip {payslip.id} - no employee")
                    continue
                    
                emp_code = employee.badge_id or ""
                emp_name = employee.get_full_name() or ""
                
                # Extract department
                dept_name = ""
                if hasattr(employee, 'employee_work_info') and employee.employee_work_info:
                    if employee.employee_work_info.department_id:
                        dept_name = employee.employee_work_info.department_id.department or ""

                # CRITICAL: Use shared calculation function to ensure consistency with table view
                calculated_values = calculate_payslip_values(payslip)
                if calculated_values is None:
                    print(f"[PAYSLIP_IMPORT_DOWNLOAD] Skipping payslip {payslip.id} - calculation returned None")
                    continue
                
                # Extract calculated values
                basic_pay = calculated_values["basic_pay"]
                housing_allowance = calculated_values["housing_allowance"]
                transport_allowance = calculated_values["transport_allowance"]
                other_allowance = calculated_values["other_allowance"]
                gross_pay = calculated_values["gross_pay"]
                paid_days = calculated_values["paid_days"]
                loss_of_pay = calculated_values["loss_of_pay"]
                overtime = calculated_values["overtime"]
                salary_advance = calculated_values["salary_advance"]
                bonus = calculated_values["bonus"]
                salary_advance_loan_recovery = calculated_values["salary_advance_loan_recovery"]
                deduction_for_export = calculated_values["deduction"]
                net_pay = calculated_values["net_pay"]
                
                # Match Excel format: Static components (A-H) then Variable components (I-P)
                export_rows.append({
                    "Employee Id": emp_code,
                    "Employee Name": emp_name,
                    "Department": dept_name,
                    "Basic Pay": basic_pay,
                    "Housing Allowance": housing_allowance,
                    "Transport Allowance": transport_allowance,
                    "Other Allowance": other_allowance,
                    "Gross Pay": gross_pay,
                    # Variable components
                    "Total Paid Days": paid_days,
                    "LOP": loss_of_pay,
                    "salary_advance_loan_recovery": salary_advance_loan_recovery,
                    "Deduction": deduction_for_export,  # Include component deductions
                    "Overtime": overtime,
                    "salary_advance": salary_advance,
                    "bonus": bonus,
                    "Net Pay": net_pay
                })
            except Exception as e:
                print(f"[PAYSLIP_IMPORT_DOWNLOAD] Error processing payslip {payslip.id if payslip else 'unknown'}: {e}")
                import traceback
                traceback.print_exc()
                continue

        if not export_rows:
            print(f"[PAYSLIP_IMPORT_DOWNLOAD] No export rows for company: {company_name}")
            continue
        
        # Mark that we're creating a sheet
        sheets_created = True
            
        # Create DataFrame for this company
        df = pd.DataFrame(export_rows)
        
        # Sanitize sheet name (Excel sheet names have restrictions)
        sheet_name = company_name[:31]  # Excel sheet name max length is 31
        sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_')
        
        # Write to Excel
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        
        # Write headers with appropriate formatting
        for col_num, col_name in enumerate(df.columns.values):
            if col_name in static_cols:
                worksheet.write(0, col_num, col_name, static_header_fmt)
            elif col_name in variable_cols:
                worksheet.write(0, col_num, col_name, variable_header_fmt)
            else:
                # Default format for any unexpected columns
                default_fmt = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3', 'border': 1})
                worksheet.write(0, col_num, col_name, default_fmt)
            worksheet.set_column(col_num, col_num, 18)
        
        # Apply row colors for data rows - preserve existing data
        # Row 0 is header, so data starts from row 1
        for row_idx in range(1, len(df) + 1):
            for col_num, col_name in enumerate(df.columns.values):
                # Get cell value from dataframe
                cell_value = df.iloc[row_idx - 1, col_num]
                
                # Apply formatting based on column type, preserving the value
                if col_name in static_cols:
                    # Static columns - green background
                    if pd.notna(cell_value):
                        worksheet.write(row_idx, col_num, cell_value, static_data_fmt)
                    else:
                        worksheet.write(row_idx, col_num, "", static_data_fmt)
                elif col_name in variable_cols:
                    # Variable columns - orange background
                    if pd.notna(cell_value):
                        worksheet.write(row_idx, col_num, cell_value, variable_data_fmt)
                    else:
                        worksheet.write(row_idx, col_num, "", variable_data_fmt)
        
        # Add labels row below data
        last_row = len(df) + 1  # +1 because header is row 0
        worksheet.write(last_row + 1, 2, "Static components for every payroll month", label_fmt_green)
        worksheet.write(last_row + 1, 8, "Variable components to be changed/entered every month", label_fmt_orange)

    # CRITICAL: Only return Excel file if at least one sheet was created
    if not sheets_created:
        writer.close()
        messages.error(request, "No payslip data found to export.")
        # Return empty response or redirect
        return redirect('payslip-info-import')
    
    writer.close()
    print(f"[PAYSLIP_IMPORT_DOWNLOAD] Successfully exported payslips to Excel")
    return response


@login_required
@permission_required("payroll.add_payslip")
def import_employees(request):
    """
    View function to handle employee import for payroll.
    """
    if request.method == 'POST':
        form = EmployeeImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            # Process the file (e.g., read data and save to database)
            # For now, just return a success message
            messages.success(request, "File uploaded successfully. Processing...")
            return redirect('payslip-info-import')
    else:
        form = EmployeeImportForm()
    return render(request, 'payroll/import_employees.html', {'form': form})


@login_required
def external_components_export(request):
    """
    Export Employee ID and Employee Name for the currently selected company.
    Company comes from navbar selection (session 'selected_company') or optional ?company_id.
    """
    company_id = request.GET.get("company_id") or request.session.get("selected_company")

    if not company_id or company_id == "all":
        employees = []
    else:
        employees = Employee.objects.filter(
            employee_work_info__company_id_id=company_id
        ).values("badge_id", "first_name", "middle_name", "last_name")

    lines = ["Employee Id,Employee Name"]
    for emp in employees:
        emp_id = emp.get("badge_id", "") or ""
        name_parts = [
            emp.get("first_name") or "",
            emp.get("middle_name") or "",
            emp.get("last_name") or "",
        ]
        emp_name = " ".join([p for p in name_parts if p]).strip()
        safe_id = '"' + emp_id.replace('"', '""') + '"'
        safe_name = '"' + emp_name.replace('"', '""') + '"'
        lines.append(f"{safe_id},{safe_name}")

    content = "\uFEFF" + "\n".join(lines)
    response = HttpResponse(content, content_type="text/csv; charset=utf-8")
    filename = f"external_components_company_{company_id or 'all'}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response